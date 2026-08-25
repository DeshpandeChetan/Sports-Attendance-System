import re
from datetime import datetime, time, timedelta
from io import BytesIO

from django.contrib import messages
from django.contrib.auth import get_user_model
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.db.models import Count, Q
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.conf import settings
from django.urls import reverse
from django.utils import timezone
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from allauth.socialaccount.models import SocialApp

from .forms import (
    AttendanceEditForm,
    DelegateForm,
    FeedbackForm,
    MembershipForm,
    ProfileForm,
    ReportFilterForm,
    SchoolForm,
    SessionFeedbackForm,
    SessionForm,
    SportForm,
    TeamForm,
    VenueForm,
)
from .email_validation import validate_christ_email
from .email_notifications import queue_notification_emails
from .models import (
    AttendanceDelegate,
    AttendanceDelegateLog,
    AttendanceEditLog,
    AttendanceRecord,
    Feedback,
    LoginAccessRequest,
    Meeting,
    Membership,
    Notification,
    School,
    Session,
    Sport,
    Team,
    UserProfile,
    Venue,
)
from .permissions import ADMIN_ROLES, TRAINER_ROLES, can_manage_team, can_schedule_session, can_take_attendance, is_admin_user, role_for, role_required

User = get_user_model()


SPORT_ICON_MAP = {
    "athletics": "bi-person-walking",
    "badminton": "bi-circle",
    "basketball": "bi-dribbble",
    "chess": "bi-puzzle",
    "cricket": "bi-trophy",
    "football": "bi-dribbble",
    "hockey": "bi-record-circle",
    "kabaddi": "bi-people",
    "soccer": "bi-dribbble",
    "swimming": "bi-water",
    "table tennis": "bi-circle",
    "tennis": "bi-circle",
    "volleyball": "bi-circle",
}
DEFAULT_SPORT_ICON = "bi-trophy"


def sport_icon_class(sport_name):
    normalized = re.sub(r"\s+", " ", str(sport_name or "").strip().lower())
    return SPORT_ICON_MAP.get(normalized, DEFAULT_SPORT_ICON)


def login_page(request):
    if request.user.is_authenticated:
        return redirect("dashboard")
    google_configured = SocialApp.objects.filter(provider="google", sites__id=1).exists()
    return render(request, "account/login.html", {"google_configured": google_configured})


@login_required
def dashboard(request):
    role = role_for(request.user)
    today = timezone.localdate()
    sessions = visible_sessions(request.user)
    trainer_stats = None
    completed_sessions = sessions.filter(attendance_submitted=True)[:8]
    upcoming_sessions = sessions.filter(start_at__date__gte=today)[:8]
    male_teams = Team.objects.filter(gender=Team.TeamGender.MALE).count()
    female_teams = Team.objects.filter(gender=Team.TeamGender.FEMALE).count()
    student_count = User.objects.filter(profile__role=UserProfile.Role.MEMBER).count()
    trainer_count = User.objects.filter(profile__role__in=[UserProfile.Role.TRAINER, UserProfile.Role.COORDINATOR]).count()
    sport_count = Sport.objects.count()
    sport_stats = Sport.objects.annotate(session_count=Count("teams__sessions")).order_by("name")
    if role in TRAINER_ROLES and not is_admin_user(request.user):
        assigned_teams = trainer_assigned_teams(request.user)
        sessions = Session.objects.filter(team__in=assigned_teams).select_related("team", "team__sport")
        completed_sessions = sessions.filter(attendance_submitted=True)[:8]
        upcoming_sessions = sessions.filter(start_at__date__gte=today, attendance_submitted=False)[:8]
        sport_stats = Sport.objects.filter(teams__in=assigned_teams).annotate(
            session_count=Count("teams__sessions", filter=Q(teams__in=assigned_teams), distinct=True)
        ).distinct().order_by("name")
        trainer_stats = {
            "sports": assigned_teams.values("sport").distinct().count(),
            "teams": assigned_teams.count(),
            "students": User.objects.filter(memberships__team__in=assigned_teams, memberships__is_active=True).distinct().count(),
            "sessions": sessions.count(),
            "completed": sessions.filter(attendance_submitted=True).count(),
            "upcoming": sessions.filter(start_at__date__gte=today, attendance_submitted=False).count(),
        }
    context = {
        "role": role,
        "dashboard_role_title": (
            "Admin"
            if is_admin_user(request.user)
            else "Trainer"
            if role in TRAINER_ROLES
            else "Student"
        ),
        "today_sessions": add_session_permissions(request.user, sessions.filter(start_at__date=today)[:8]),
        "today_meetings": visible_meetings(request.user).prefetch_related("sports", "teams", "trainers", "participants").select_related("scheduled_by").filter(meeting_date=today)[:8],
        "upcoming_sessions": add_session_permissions(request.user, upcoming_sessions),
        "completed_sessions": add_session_permissions(request.user, completed_sessions),
        "attendance_percent": attendance_percentage(request.user),
        "unread_feedback": Feedback.objects.filter(receiver=request.user, is_read=False).count(),
        "sport_stats": sport_stats,
        "trainer_stats": trainer_stats,
        "admin_stats": {
            "sports": sport_count,
            "students": student_count,
            "trainers": trainer_count,
            "male_teams": male_teams,
            "female_teams": female_teams,
            "teams": male_teams + female_teams,
        },
        "dashboard_chart_data": {
            "teamGender": {"labels": ["Male Teams", "Female Teams"], "data": [male_teams, female_teams]},
            "people": {"labels": ["Students", "Trainers"], "data": [student_count, trainer_count]},
            "inventory": {"labels": ["Sports", "Teams"], "data": [sport_count, male_teams + female_teams]},
        },
    }
    return render(request, "core/dashboard.html", context)


@login_required
@role_required(*ADMIN_ROLES)
def analytics(request):
    today = timezone.localdate()
    male_teams = Team.objects.filter(gender=Team.TeamGender.MALE).count()
    female_teams = Team.objects.filter(gender=Team.TeamGender.FEMALE).count()
    student_count = User.objects.filter(profile__role=UserProfile.Role.MEMBER).count()
    trainer_count = User.objects.filter(profile__role__in=[UserProfile.Role.TRAINER, UserProfile.Role.COORDINATOR]).count()
    sport_count = Sport.objects.count()
    team_count = male_teams + female_teams
    total_attendance = AttendanceRecord.objects.count()
    attended_attendance = AttendanceRecord.objects.filter(status__in=[
        AttendanceRecord.Status.PRESENT,
        AttendanceRecord.Status.LATE,
        AttendanceRecord.Status.EARLY_EXIT,
    ]).count()
    absent_attendance = AttendanceRecord.objects.filter(status=AttendanceRecord.Status.ABSENT).count()
    attendance_percent = round((attended_attendance / total_attendance) * 100, 1) if total_attendance else None
    total_sessions = Session.objects.count()
    completed_sessions = Session.objects.filter(attendance_submitted=True).count()
    upcoming_sessions = Session.objects.filter(start_at__date__gte=today, attendance_submitted=False).count()
    assigned_trainers = User.objects.filter(
        profile__role__in=[UserProfile.Role.TRAINER, UserProfile.Role.COORDINATOR],
        coordinated_teams__isnull=False,
    ).distinct().count()
    trainer_rows = User.objects.filter(
        Q(profile__role__in=[UserProfile.Role.TRAINER, UserProfile.Role.COORDINATOR])
        | Q(coordinated_teams__isnull=False)
    ).annotate(
        assigned_team_count=Count("coordinated_teams", distinct=True),
        session_count=Count("coordinated_teams__sessions", distinct=True),
        completed_session_count=Count("coordinated_teams__sessions", filter=Q(coordinated_teams__sessions__attendance_submitted=True), distinct=True),
        upcoming_session_count=Count("coordinated_teams__sessions", filter=Q(coordinated_teams__sessions__start_at__date__gte=today, coordinated_teams__sessions__attendance_submitted=False), distinct=True),
    ).distinct().order_by("first_name", "last_name", "email")
    chart_data = {
        "teamGender": {"labels": ["Male Teams", "Female Teams"], "data": [male_teams, female_teams]},
        "people": {"labels": ["Students", "Trainers"], "data": [student_count, trainer_count]},
        "inventory": {"labels": ["Sports", "Teams"], "data": [sport_count, team_count]},
        "attendance": {"labels": ["Attended", "Absent"], "data": [attended_attendance, absent_attendance]},
        "sessions": {"labels": ["Completed", "Upcoming", "Other"], "data": [completed_sessions, upcoming_sessions, max(total_sessions - completed_sessions - upcoming_sessions, 0)]},
        "trainers": {"labels": ["Assigned", "Unassigned"], "data": [assigned_trainers, max(trainer_count - assigned_trainers, 0)]},
    }
    return render(request, "core/analytics.html", {
        "dashboard_chart_data": chart_data,
        "analytics_stats": {
            "attendance_percent": attendance_percent,
            "attendance_records": total_attendance,
            "attended_records": attended_attendance,
            "absent_records": absent_attendance,
            "sessions": total_sessions,
            "completed_sessions": completed_sessions,
            "upcoming_sessions": upcoming_sessions,
            "trainers": trainer_count,
            "assigned_trainers": assigned_trainers,
        },
        "trainer_rows": trainer_rows,
    })


def visible_sessions(user):
    qs = Session.objects.select_related("team", "team__sport")
    if is_admin_user(user):
        return qs
    return qs.filter(
        Q(team__memberships__user=user, team__memberships__is_active=True)
        | Q(team__captain=user)
        | Q(team__vice_captain=user)
        | Q(team__coordinator=user)
        | Q(delegates__assigned_to=user)
    ).distinct()


def visible_meetings(user):
    return Meeting.objects.filter(Q(participants=user) | Q(scheduled_by=user)).distinct()


def add_session_permissions(user, sessions):
    now = timezone.now()
    session_list = list(sessions)
    for item in session_list:
        item.can_manage_for_user = can_manage_team(user, item.team)
        item.can_take_for_user = can_take_attendance(user, item)
        item.has_active_attendance_lock = attendance_lock_active(item, now)
        item.sport_icon_class = sport_icon_class(item.team.sport.name)
        item.scheduled_by_name = user_label(item.scheduled_by) if item.scheduled_by else "-"
    return session_list


def admin_users():
    return User.objects.filter(Q(is_superuser=True) | Q(profile__role__in=ADMIN_ROLES), is_active=True).distinct()


def users_for_team(team):
    users = User.objects.filter(
        Q(memberships__team=team, memberships__is_active=True)
        | Q(pk=team.captain_id)
        | Q(pk=team.vice_captain_id)
        | Q(pk=team.coordinator_id),
        is_active=True,
    ).distinct()
    return users


def trainer_assigned_teams(user):
    if is_admin_user(user):
        return Team.objects.filter(is_active=True)
    if role_for(user) in TRAINER_ROLES:
        return Team.objects.filter(is_active=True, coordinator=user)
    return Team.objects.none()


def can_manage_student_membership(user, membership):
    if is_admin_user(user):
        return True
    return bool(membership and can_manage_team(user, membership.team))


def can_view_student_user(user, student):
    if is_admin_user(user):
        return True
    if role_for(user) in TRAINER_ROLES:
        return Membership.objects.filter(user=student, team__coordinator=user).exists()
    return False


def create_notifications(users, title, message, actor=None, target_url="", sport=None, team=None, session=None, include_actor=False, email_type=None, email_context=None):
    notifications = []
    recipients = []
    seen = set()
    actor_id = getattr(actor, "pk", None)
    for user in users:
        if not user or user.pk in seen or (user.pk == actor_id and not include_actor):
            continue
        seen.add(user.pk)
        recipients.append(user)
        notifications.append(Notification(
            user=user,
            actor=actor if getattr(actor, "is_authenticated", False) else None,
            title=title,
            message=message,
            target_url=target_url,
            sport=sport,
            team=team,
            session=session,
        ))
    if notifications:
        Notification.objects.bulk_create(notifications)
        if email_type:
            context = {"title": title, "message": message, **(email_context or {})}
            # Email is best-effort and internally catches every delivery failure.
            queue_notification_emails(recipients, email_type, context)


def notify_team_membership_change(actor, users, team, change):
    if not users:
        return
    labels = {
        "added": ("Team membership added", "added to"),
        "removed": ("Team membership removed", "removed from"),
        "activated": ("Team membership activated", "activated for"),
        "deactivated": ("Team membership deactivated", "deactivated for"),
    }
    title, verb = labels[change]
    message = f"You have been {verb} {team.sport.name} - {team.get_gender_display()} {team.name}."
    create_notifications(
        users,
        title,
        message,
        actor=actor,
        target_url=reverse("teams"),
        sport=team.sport,
        team=team,
        include_actor=True,
        email_type="team_membership",
        email_context={"details": [("Sport", team.sport.name), ("Team", f"{team.get_gender_display()} {team.name}"), ("Changed By", user_label(actor))]},
    )


def notify_team_role_change(actor, user, team, role_name, assigned=True):
    if not user:
        return
    action = "assigned as" if assigned else "removed as"
    title = f"{role_name} {'assigned' if assigned else 'removed'}"
    message = f"You have been {action} {role_name} for {team.sport.name} - {team.get_gender_display()} {team.name}."
    create_notifications(
        [user],
        title,
        message,
        actor=actor,
        target_url=reverse("teams"),
        sport=team.sport,
        team=team,
        include_actor=True,
        email_type=f"{role_name.lower().replace(' ', '_')}_{'assigned' if assigned else 'removed'}",
        email_context={"details": [("Sport", team.sport.name), ("Team", f"{team.get_gender_display()} {team.name}"), ("Role", role_name), ("Updated By", user_label(actor))]},
    )


def notify_practice_session(session, title, message, actor, teams=None, include_session=True, email_type="practice_updated"):
    notification_teams = teams or [session.team]
    for team in notification_teams:
        create_notifications(
            users_for_team(team),
            title,
            message,
            actor=actor,
            target_url=reverse("sessions"),
            sport=team.sport,
            team=team,
            session=session if include_session else None,
            email_type=email_type,
            email_context={
                "details": [
                    ("Sport", team.sport.name),
                    ("Team", f"{team.get_gender_display()} {team.name}"),
                    ("Date", timezone.localtime(session.start_at).strftime("%d %B %Y")),
                    ("Time / Schedule", session_time_schedule_label(session)),
                    ("Venue", session.venue),
                    ("Scheduled / Updated By", user_label(actor)),
                ]
            },
        )


def notify_common_action(actor, title, message, target_url, email_type=None, email_context=None):
    create_notifications(User.objects.filter(is_active=True), title, message, actor=actor, target_url=target_url, email_type=email_type, email_context=email_context)


def user_label(user):
    return user.get_full_name() or user.email or user.username


def profile_for_user(user):
    profile, _ = UserProfile.objects.get_or_create(user=user)
    return profile


def set_user_account_status(user, status, note=""):
    profile = profile_for_user(user)
    profile.account_status = status
    profile.account_status_note = "" if status == UserProfile.AccountStatus.ACTIVE else str(note or "").strip()
    profile.save(update_fields=["account_status", "account_status_note", "updated_at"])
    should_be_active = status == UserProfile.AccountStatus.ACTIVE
    if user.is_active != should_be_active:
        user.is_active = should_be_active
        user.save(update_fields=["is_active"])


def session_time_schedule_label(session):
    return f"{timezone.localtime(session.start_at):%I:%M %p} - {timezone.localtime(session.end_at):%I:%M %p}{' (Full Day)' if session.full_day else ''} ({session.get_schedule_slot_display()})"


def user_account_status(user):
    profile = profile_for_user(user)
    if profile.account_status != UserProfile.AccountStatus.ACTIVE:
        return profile.account_status
    if not user.is_active:
        return UserProfile.AccountStatus.DEACTIVATED
    return UserProfile.AccountStatus.ACTIVE


def attendance_lock_active(session, now=None):
    now = now or timezone.now()
    return bool(session.attendance_started_by and session.attendance_lock_expires_at and session.attendance_lock_expires_at > now)


def attendance_percentage(user):
    total = AttendanceRecord.objects.filter(member=user).count()
    if total == 0:
        return None
    positive = AttendanceRecord.objects.filter(member=user, status__in=["PRESENT", "LATE", "EARLY_EXIT"]).count()
    return round((positive / total) * 100, 1)


def role_label(user):
    if not user or not user.is_authenticated:
        return ""
    if user.is_superuser:
        return "Super Admin"
    return user.profile.get_role_display()


def meeting_participant_role_label(user, teams, organizer=None):
    if organizer and user.pk == organizer.pk:
        return role_label(user)
    roles = []
    for team in teams:
        if team.captain_id == user.pk:
            roles.append("Captain")
        if team.vice_captain_id == user.pk:
            roles.append("Vice Captain")
        if team.coordinator_id == user.pk:
            roles.append("Trainer")
    if roles:
        return ", ".join(dict.fromkeys(roles))
    return role_label(user)


def meeting_participant_sports_label(user, teams, organizer=None):
    selected_sport_names = sorted({team.sport.name for team in teams})
    if organizer and user.pk == organizer.pk:
        return ", ".join(selected_sport_names) or "-"
    sport_names = set()
    for team in teams:
        if user.pk in {team.captain_id, team.vice_captain_id, team.coordinator_id}:
            sport_names.add(team.sport.name)
    sport_names.update(
        Membership.objects.filter(
            user=user,
            is_active=True,
            team__in=teams,
        ).values_list("team__sport__name", flat=True)
    )
    sport_names.update(
        Team.objects.filter(
            coordinator=user,
            sport__name__in=selected_sport_names,
            is_active=True,
        ).values_list("sport__name", flat=True)
    )
    return ", ".join(sorted(sport_names)) or "-"


def split_full_name(full_name):
    parts = (full_name or "").strip().split(" ", 1)
    return parts[0] if parts else "", parts[1] if len(parts) > 1 else ""


def normalize_header(value):
    return re.sub(r"[^a-z0-9]+", "_", str(value or "").strip().lower()).strip("_")


def truthy_cell(value, default=True):
    if value in (None, ""):
        return default
    return str(value).strip().lower() in {"1", "yes", "y", "true", "active"}


def parse_time_cell(value, fallback):
    if value in (None, ""):
        return fallback
    if isinstance(value, datetime):
        return value.time().replace(second=0, microsecond=0)
    if isinstance(value, time):
        return value.replace(second=0, microsecond=0)
    text = str(value).strip()
    for fmt in ("%H:%M", "%I:%M %p", "%I:%M%p"):
        try:
            return datetime.strptime(text, fmt).time()
        except ValueError:
            continue
    raise ValueError("Times must use HH:MM format.")


BULK_UPLOAD_TEMPLATES = {
    "sports": {
        "filename": "sports_bulk_upload_sample.xlsx",
        "headers": ["Sport Name", "Description"],
        "rows": [["Basketball", "Indoor and outdoor basketball training"]],
    },
    "venues": {
        "filename": "venues_bulk_upload_sample.xlsx",
        "headers": ["Venue Name", "Location", "Status"],
        "rows": [["Indoor Stadium", "Main sports block", "Active"]],
    },
    "schools": {
        "filename": "schools_bulk_upload_sample.xlsx",
        "headers": ["School Name", "Description", "Status"],
        "rows": [["School of Sciences", "Science and data science programmes", "Active"]],
    },
    "teams": {
        "filename": "teams_bulk_upload_sample.xlsx",
        "headers": ["Sport", "Team Name", "Team Type", "Gender", "Captain Email", "Vice Captain Email", "Trainer Email", "Status"],
        "rows": [["Basketball", "Men's Team A", "University", "Male", "", "", "coach@christuniversity.in", "Active"]],
    },
    "students": {
        "filename": "students_bulk_upload_sample.xlsx",
        "headers": ["Student Name", "Student Email", "Mobile Number", "Reg No", "School", "Department", "Gender", "Sport", "Team Type", "Team", "Status"],
        "rows": [["Rahul Sharma", "rahul.sharma@christuniversity.in", "9876543210", "22104321", "School of Sciences", "MSDS", "Male", "Basketball", "University", "Men's Team A", "Active"]],
    },
    "trainers": {
        "filename": "trainers_bulk_upload_sample.xlsx",
        "headers": ["Trainer Name", "Trainer Email", "Mobile Number", "Reg No", "Teams"],
        "rows": [["Coach Ramesh", "coach.ramesh@christuniversity.in", "9876543211", "TRN001", "Basketball - Male Men's Team A"]],
    },
    "sessions": {
        "filename": "sessions_bulk_upload_sample.xlsx",
        "headers": ["Team", "Title", "Start Date", "End Date", "From Time", "To Time", "Schedule", "Full Day", "Venue", "Other Venue", "Notes"],
        "rows": [["Basketball - Male Men's Team A", "Morning Tactical Drill", "2026-07-20", "2026-07-20", "06:30", "08:30", "Morning", "No", "Indoor Stadium", "", "Fitness and drills"]],
    },
}


def read_bulk_upload_rows(uploaded_file):
    if not uploaded_file:
        raise ValueError("Please choose an Excel file to upload.")
    workbook = load_workbook(uploaded_file, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [normalize_header(cell) for cell in rows[0]]
    data = []
    for row in rows[1:]:
        if not any(cell not in (None, "") for cell in row):
            continue
        data.append({headers[index]: row[index] if index < len(row) else "" for index in range(len(headers))})
    return data


def add_bulk_upload_messages(request, created, errors):
    if created:
        messages.success(request, f"Bulk upload completed. {created} row(s) saved.")
    if errors:
        detail = " ".join(errors[:5])
        extra = f" {len(errors) - 5} more error(s)." if len(errors) > 5 else ""
        messages.error(request, f"{len(errors)} row(s) could not be uploaded. {detail}{extra}")


def missing_required(row, fields):
    return [label for key, label in fields if not str(row.get(key) or "").strip()]


def find_user_by_email(email):
    if not email:
        return None
    return User.objects.filter(email__iexact=str(email).strip()).first()


def is_student_leader_candidate(user):
    if not user:
        return True
    role = getattr(getattr(user, "profile", None), "role", None)
    return role in {UserProfile.Role.MEMBER, UserProfile.Role.CAPTAIN, UserProfile.Role.VICE_CAPTAIN}


def normalize_choice_label(value):
    return str(value or "").strip().lower()


def find_team_type(value):
    value = normalize_choice_label(value)
    if not value:
        return None
    for stored_value, label in Team.TeamType.choices:
        if value in {stored_value.lower(), label.lower()}:
            return stored_value
    return None


def normalize_gender(value):
    value = normalize_choice_label(value)
    if value in {"male", "m"}:
        return "Male"
    if value in {"female", "f"}:
        return "Female"
    return ""


def find_team_by_label(label, sport=None, team_type=None):
    label = str(label or "").strip().lower()
    if not label:
        return None
    teams = Team.objects.select_related("sport")
    sport = str(sport or "").strip()
    resolved_team_type = find_team_type(team_type)
    if sport:
        teams = teams.filter(sport__name__iexact=sport)
    if resolved_team_type:
        teams = teams.filter(team_type=resolved_team_type)
    for team in teams:
        labels = {
            team.name.lower(),
            str(team).lower(),
            f"{team.sport.name} - {team.get_gender_display()} {team.name}".lower(),
            f"{team.sport.name} - {team.get_team_type_display()} - {team.get_gender_display()} {team.name}".lower(),
        }
        if label in labels:
            return team
    return None


def membership_conflict_message(teams):
    seen = {}
    for team in teams:
        key = (team.sport_id, team.gender)
        existing = seen.get(key)
        if existing and existing.pk != team.pk:
            return (
                "A student can belong to only one Team Category within the same Sport and Gender. "
                f"Conflict: {existing} and {team}."
            )
        seen[key] = team
    return ""


def validate_student_team_memberships(student, selected_teams, replace_team_ids=None):
    replace_team_ids = set(replace_team_ids or [])
    existing_teams = Team.objects.filter(memberships__user=student).exclude(pk__in=replace_team_ids).select_related("sport")
    return membership_conflict_message(list(existing_teams) + list(selected_teams))


def parse_int_ids(values):
    return [int(value) for value in values if str(value).isdigit()]


def build_meeting_preview(post_data, organizer=None):
    title = (post_data.get("title") or "").strip()
    meeting_date = post_data.get("meeting_date")
    start_time = post_data.get("start_time")
    end_time = post_data.get("end_time")
    venue_choice = (post_data.get("venue") or "").strip()
    other_venue = (post_data.get("other_venue") or "").strip()
    venue = f"Other - {other_venue}" if venue_choice == "OTHER" and other_venue else venue_choice
    agenda = (post_data.get("agenda") or "").strip()
    sport_ids = parse_int_ids(post_data.getlist("sports"))
    team_ids = parse_int_ids(post_data.getlist("teams"))
    trainer_ids = parse_int_ids(post_data.getlist("trainers"))
    errors = []
    if not title:
        errors.append("Meeting Title is required.")
    if not meeting_date:
        errors.append("Date is required.")
    if not start_time:
        errors.append("Start Time is required.")
    if not end_time:
        errors.append("End Time is required.")
    if start_time and end_time and start_time >= end_time:
        errors.append("End Time must be after Start Time.")
    if not venue_choice:
        errors.append("Venue is required.")
    if venue_choice == "OTHER" and not other_venue:
        errors.append("Specific Venue is required when Venue is Other.")
    if not sport_ids:
        errors.append("Select at least one Sport.")
    if not team_ids:
        errors.append("Select at least one Team.")

    sports = Sport.objects.filter(pk__in=sport_ids, is_active=True).order_by("name")
    teams = Team.objects.filter(pk__in=team_ids, is_active=True, sport_id__in=sport_ids).select_related("sport").order_by("sport__name", "gender", "name")
    trainers = User.objects.filter(
        pk__in=trainer_ids,
        is_active=True,
        profile__role__in=[UserProfile.Role.TRAINER, UserProfile.Role.COORDINATOR],
    ).select_related("profile").order_by("first_name", "last_name", "email")
    if len(sport_ids) != sports.count():
        errors.append("One or more selected Sports are invalid.")
    if len(team_ids) != teams.count():
        errors.append("One or more selected Teams do not belong to the selected Sports.")
    if len(trainer_ids) != trainers.count():
        errors.append("One or more selected Trainers are invalid.")

    participants_by_id = {trainer.pk: trainer for trainer in trainers}
    if organizer:
        participants_by_id[organizer.pk] = organizer
    team_modes = {}
    for team in teams:
        mode = post_data.get(f"team_participants_{team.pk}", "LEADS")
        if mode not in {"LEADS", "ALL"}:
            mode = "LEADS"
        team_modes[str(team.pk)] = mode
        if mode == "ALL":
            team_users = User.objects.filter(memberships__team=team, memberships__is_active=True, is_active=True)
        else:
            lead_ids = [value for value in [team.captain_id, team.vice_captain_id] if value]
            team_users = User.objects.filter(pk__in=lead_ids, is_active=True)
        for participant in team_users.select_related("profile"):
            participants_by_id[participant.pk] = participant
    participants = sorted(participants_by_id.values(), key=lambda item: (user_label(item).lower(), item.pk))
    if not participants:
        errors.append("No participants were found for the selected Trainers/Teams.")
    participant_details = [
        {
            "name": user_label(participant),
            "role": meeting_participant_role_label(participant, teams, organizer=organizer),
            "sports": meeting_participant_sports_label(participant, teams, organizer=organizer),
        }
        for participant in participants
    ]
    try:
        display_date = datetime.strptime(meeting_date, "%Y-%m-%d").strftime("%d/%m/%Y") if meeting_date else ""
    except ValueError:
        display_date = meeting_date or ""
    try:
        display_start_time = datetime.strptime(start_time, "%H:%M").strftime("%I:%M %p").lower() if start_time else ""
        display_end_time = datetime.strptime(end_time, "%H:%M").strftime("%I:%M %p").lower() if end_time else ""
    except ValueError:
        display_start_time = start_time or ""
        display_end_time = end_time or ""

    return {
        "errors": errors,
        "data": {
            "title": title,
            "meeting_date": meeting_date,
            "start_time": start_time,
            "end_time": end_time,
            "venue_choice": venue_choice,
            "venue": venue,
            "other_venue": other_venue,
            "agenda": agenda,
            "sport_ids": [sport.pk for sport in sports],
            "team_ids": [team.pk for team in teams],
            "trainer_ids": [trainer.pk for trainer in trainers],
            "participant_ids": [participant.pk for participant in participants],
            "team_modes": team_modes,
        },
        "preview": {
            "title": title,
            "meeting_date": meeting_date,
            "display_date": display_date,
            "start_time": start_time,
            "end_time": end_time,
            "display_start_time": display_start_time,
            "display_end_time": display_end_time,
            "venue": venue,
            "agenda": agenda,
            "sports": [sport.name for sport in sports],
            "teams": [str(team) for team in teams],
            "trainers": [user_label(trainer) for trainer in trainers],
            "participants": participant_details,
            "team_modes": {str(team.pk): "Complete Team" if team_modes.get(str(team.pk)) == "ALL" else "Captain and Vice Captain Only" for team in teams},
        },
    }


def meeting_display_date(value):
    if not value:
        return ""
    if hasattr(value, "strftime"):
        return value.strftime("%d/%m/%Y")
    try:
        return datetime.strptime(str(value), "%Y-%m-%d").strftime("%d/%m/%Y")
    except ValueError:
        return str(value)


def meeting_display_time(value):
    if not value:
        return ""
    if hasattr(value, "strftime"):
        return value.strftime("%I:%M %p").lower()
    try:
        return datetime.strptime(str(value), "%H:%M").strftime("%I:%M %p").lower()
    except ValueError:
        return str(value)


def meeting_email_details(title, meeting_date, start_time, end_time, venue, agenda, actor, actor_label="Updated By"):
    return [
        ("Title", title),
        ("Date", meeting_display_date(meeting_date)),
        ("Time", f"{meeting_display_time(start_time)} - {meeting_display_time(end_time)}"),
        ("Venue", venue),
        ("Agenda", agenda),
        (actor_label, user_label(actor)),
    ]


def saved_meeting_schedule_tuple(meeting):
    return (
        meeting.meeting_date.isoformat() if meeting.meeting_date else "",
        meeting.start_time.strftime("%H:%M") if meeting.start_time else "",
        meeting.end_time.strftime("%H:%M") if meeting.end_time else "",
        meeting.venue,
    )


def resolve_school(value=None, school_id=None, create=False):
    if school_id and str(school_id).isdigit():
        return School.objects.filter(pk=int(school_id)).first()
    school_name = str(value or "").strip()
    if not school_name:
        return None
    if create:
        school, _ = School.objects.get_or_create(name=school_name, defaults={"is_active": True})
        return school
    return School.objects.filter(name__iexact=school_name).first()


def save_student_record(full_name, email, school=None, department="", phone="", register_no="", gender="", existing_user=None):
    email = str(email or "").strip().lower()
    phone = str(phone or "").strip()
    register_no = str(register_no or "").strip()
    if not email:
        raise ValueError("Student email is required.")
    email = validate_christ_email(email)
    if phone and not re.fullmatch(r"\d{10}", phone):
        raise ValueError("Mobile number must be exactly 10 digits.")
    user = existing_user or User.objects.filter(email__iexact=email).first()
    validate_unique_register_no(register_no, user)
    if user is None:
        user = User(username=email, email=email)
        user.set_unusable_password()
    if full_name:
        user.first_name, user.last_name = split_full_name(full_name)
    user.email = email
    user.username = email
    user.save()
    profile, _ = UserProfile.objects.get_or_create(user=user)
    profile.role = UserProfile.Role.MEMBER
    profile.school = school
    profile.department = str(department or "").strip()
    profile.phone = phone
    profile.register_no = register_no
    profile.gender = normalize_gender(gender)
    profile.save(update_fields=["role", "school", "department", "phone", "register_no", "gender", "updated_at"])
    return user


def validate_unique_register_no(register_no, user=None):
    register_no = str(register_no or "").strip()
    if not register_no:
        return
    existing = UserProfile.objects.filter(register_no__iexact=register_no)
    if user and user.pk:
        existing = existing.exclude(user=user)
    if existing.exists():
        raise ValueError("This registration number already exists")


def build_student_from_post(request, existing_user=None):
    full_name = request.POST.get("student_name", "").strip()
    email = request.POST.get("student_email", "").strip().lower()
    school = resolve_school(school_id=request.POST.get("school"))
    department = request.POST.get("department", "").strip()
    phone = request.POST.get("mobile_number", "").strip()
    register_no = request.POST.get("register_no", "").strip()
    gender = request.POST.get("gender", "").strip()
    return save_student_record(full_name, email, school, department, phone, register_no, gender, existing_user=existing_user)


def create_student_from_access_request(access_request):
    email = validate_christ_email(access_request.email)
    user = User.objects.filter(email__iexact=email).first()
    if user is None:
        user = User(username=email, email=email)
        user.set_unusable_password()
    user.email = email
    user.username = email
    if access_request.full_name:
        parts = access_request.full_name.split(" ", 1)
        user.first_name = parts[0]
        user.last_name = parts[1] if len(parts) > 1 else ""
    user.save()
    profile, _ = UserProfile.objects.get_or_create(user=user)
    profile.role = UserProfile.Role.MEMBER
    profile.save(update_fields=["role", "updated_at"])
    return user


@login_required
def bulk_upload_sample(request, kind):
    template = BULK_UPLOAD_TEMPLATES.get(kind)
    if not template:
        messages.error(request, "Unknown bulk upload sample.")
        return redirect("dashboard")
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sample"
    sheet.append(template["headers"])
    for row in template["rows"]:
        sheet.append(row)
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    response = HttpResponse(output.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="{template["filename"]}"'
    return response


def import_sports_from_file(uploaded_file):
    created = 0
    errors = []
    for index, row in enumerate(read_bulk_upload_rows(uploaded_file), start=2):
        name = str(row.get("sport_name") or row.get("name") or "").strip()
        if not name:
            errors.append(f"Row {index}: Sport Name is required.")
            continue
        Sport.objects.update_or_create(name=name, defaults={"description": str(row.get("description") or "").strip(), "is_active": True})
        created += 1
    return created, errors


def import_venues_from_file(uploaded_file):
    created = 0
    errors = []
    for index, row in enumerate(read_bulk_upload_rows(uploaded_file), start=2):
        name = str(row.get("venue_name") or row.get("name") or "").strip()
        if not name:
            errors.append(f"Row {index}: Venue Name is required.")
            continue
        Venue.objects.update_or_create(
            name=name,
            defaults={"location": str(row.get("location") or "").strip(), "is_active": truthy_cell(row.get("status") or row.get("active"), True)},
        )
        created += 1
    return created, errors


def import_schools_from_file(uploaded_file):
    created = 0
    errors = []
    for index, row in enumerate(read_bulk_upload_rows(uploaded_file), start=2):
        name = str(row.get("school_name") or row.get("name") or "").strip()
        if not name:
            errors.append(f"Row {index}: School Name is required.")
            continue
        School.objects.update_or_create(
            name=name,
            defaults={"description": str(row.get("description") or "").strip(), "is_active": truthy_cell(row.get("status") or row.get("active"), True)},
        )
        created += 1
    return created, errors


def import_teams_from_file(uploaded_file):
    created = 0
    errors = []
    type_map = {label.lower(): value for value, label in Team.TeamType.choices}
    gender_map = {label.lower(): value for value, label in Team.TeamGender.choices}
    for index, row in enumerate(read_bulk_upload_rows(uploaded_file), start=2):
        sport_name = str(row.get("sport") or "").strip()
        team_name = str(row.get("team_name") or row.get("name") or "").strip()
        missing = []
        if not sport_name:
            missing.append("Sport")
        if not team_name:
            missing.append("Team Name")
        if missing:
            errors.append(f"Row {index}: {', '.join(missing)} required.")
            continue
        sport, _ = Sport.objects.get_or_create(name=sport_name, defaults={"is_active": True})
        team_type = type_map.get(str(row.get("team_type") or "University").strip().lower(), Team.TeamType.UNIVERSITY)
        gender = gender_map.get(str(row.get("gender") or "Male").strip().lower(), Team.TeamGender.MALE)
        captain = find_user_by_email(row.get("captain_email"))
        vice_captain = find_user_by_email(row.get("vice_captain_email"))
        current_team = Team.objects.filter(sport=sport, name=team_name, team_type=team_type, gender=gender).first()
        if not is_student_leader_candidate(captain):
            errors.append(f"Row {index}: Captain must be a student, not a trainer/admin.")
            continue
        if not is_student_leader_candidate(vice_captain):
            errors.append(f"Row {index}: Vice Captain must be a student, not a trainer/admin.")
            continue
        if captain and vice_captain and captain == vice_captain:
            errors.append(f"Row {index}: Captain and Vice Captain cannot be the same student.")
            continue
        leader_conflict = False
        for label, leader in (("Captain", captain), ("Vice Captain", vice_captain)):
            if not leader:
                continue
            existing = Team.objects.filter(Q(captain=leader) | Q(vice_captain=leader))
            if current_team:
                existing = existing.exclude(pk=current_team.pk)
            if existing.exists():
                errors.append(f"Row {index}: {label} is already Captain or Vice Captain of another team.")
                leader_conflict = True
        if leader_conflict:
            continue
        Team.objects.update_or_create(
            sport=sport,
            name=team_name,
            team_type=team_type,
            gender=gender,
            defaults={
                "captain": captain,
                "vice_captain": vice_captain,
                "coordinator": find_user_by_email(row.get("trainer_email")),
                "is_active": truthy_cell(row.get("status") or row.get("active"), True),
            },
        )
        created += 1
    return created, errors


def import_students_from_file(uploaded_file, user=None):
    created = 0
    errors = []
    for index, row in enumerate(read_bulk_upload_rows(uploaded_file), start=2):
        missing = missing_required(row, [("student_name", "Student Name"), ("student_email", "Student Email")])
        if missing:
            errors.append(f"Row {index}: {', '.join(missing)} required.")
            continue
        sport_name = str(row.get("sport") or "").strip()
        team_type = str(row.get("team_type") or "").strip()
        team = find_team_by_label(row.get("team"), sport=sport_name, team_type=team_type)
        if row.get("team") and not team:
            errors.append(f"Row {index}: Team was not found for the selected Sport and Team Type.")
            continue
        if user and not is_admin_user(user):
            if not team:
                errors.append(f"Row {index}: Team is required for trainer uploads.")
                continue
            if not can_manage_team(user, team):
                errors.append(f"Row {index}: You cannot add students to this team.")
                continue
        try:
            school = resolve_school(row.get("school"), create=True)
            student = save_student_record(
                row.get("student_name"),
                row.get("student_email"),
                school,
                row.get("department") or row.get("class"),
                row.get("mobile_number"),
                row.get("reg_no") or row.get("register_no"),
                row.get("gender"),
            )
            if team:
                conflict = validate_student_team_memberships(student, [team], replace_team_ids=[team.pk])
                if conflict:
                    errors.append(f"Row {index}: {conflict}")
                    continue
                membership, membership_created = Membership.objects.get_or_create(user=student, team=team)
                was_active = membership.is_active
                new_active = truthy_cell(row.get("status") or row.get("active"), True)
                membership.is_active = new_active
                membership.save(update_fields=["is_active", "updated_at"])
                if new_active and membership_created:
                    notify_team_membership_change(user, [student], team, "added")
                elif new_active and not was_active:
                    notify_team_membership_change(user, [student], team, "activated")
                elif was_active and not new_active:
                    notify_team_membership_change(user, [student], team, "deactivated")
            created += 1
        except ValueError as exc:
            errors.append(f"Row {index}: {exc}")
    return created, errors


def import_trainers_from_file(uploaded_file):
    created = 0
    errors = []
    for index, row in enumerate(read_bulk_upload_rows(uploaded_file), start=2):
        missing = missing_required(row, [("trainer_name", "Trainer Name"), ("trainer_email", "Trainer Email")])
        if missing:
            errors.append(f"Row {index}: {', '.join(missing)} required.")
            continue
        email = str(row.get("trainer_email") or "").strip().lower()
        phone = str(row.get("mobile_number") or "").strip()
        try:
            email = validate_christ_email(email)
        except ValueError as exc:
            errors.append(f"Row {index}: {exc}")
            continue
        if phone and not re.fullmatch(r"\d{10}", phone):
            errors.append(f"Row {index}: Mobile number must be exactly 10 digits.")
            continue
        trainer = User.objects.filter(email__iexact=email).first()
        try:
            validate_unique_register_no(row.get("reg_no") or row.get("register_no"), trainer)
        except ValueError as exc:
            errors.append(f"Row {index}: {exc}")
            continue
        if trainer is None:
            trainer = User(username=email, email=email)
            trainer.set_unusable_password()
        full_name = row.get("trainer_name")
        if full_name:
            trainer.first_name, trainer.last_name = split_full_name(full_name)
        trainer.email = email
        trainer.username = email
        trainer.save()
        profile, _ = UserProfile.objects.get_or_create(user=trainer)
        profile.role = UserProfile.Role.TRAINER
        profile.phone = phone
        profile.register_no = str(row.get("reg_no") or row.get("register_no") or "").strip()
        profile.save(update_fields=["role", "phone", "register_no", "updated_at"])
        teams = [find_team_by_label(label) for label in str(row.get("teams") or "").split(",")]
        team_ids = [team.pk for team in teams if team]
        Team.objects.filter(coordinator=trainer).exclude(pk__in=team_ids).update(coordinator=None)
        Team.objects.filter(pk__in=team_ids).update(coordinator=trainer)
        created += 1
    return created, errors


def import_sessions_from_file(uploaded_file, user):
    created = 0
    errors = []
    slot_map = {"morning": Session.ScheduleSlot.MORNING, "evening": Session.ScheduleSlot.EVENING}
    for index, row in enumerate(read_bulk_upload_rows(uploaded_file), start=2):
        missing = missing_required(row, [("team", "Team"), ("start_date", "Start Date"), ("end_date", "End Date"), ("schedule", "Schedule"), ("venue", "Venue")])
        if missing:
            errors.append(f"Row {index}: {', '.join(missing)} required.")
            continue
        team = find_team_by_label(row.get("team"))
        if not team:
            errors.append(f"Row {index}: Team was not found.")
            continue
        if not can_schedule_session(user, team):
            errors.append(f"Row {index}: You cannot schedule for this team.")
            continue
        try:
            start_date = row.get("start_date")
            end_date = row.get("end_date") or start_date
            if isinstance(start_date, datetime):
                start_date = start_date.date()
            if isinstance(end_date, datetime):
                end_date = end_date.date()
            if isinstance(start_date, str):
                start_date = datetime.strptime(start_date.strip(), "%Y-%m-%d").date()
            if isinstance(end_date, str):
                end_date = datetime.strptime(end_date.strip(), "%Y-%m-%d").date()
        except ValueError:
            errors.append(f"Row {index}: Dates must use YYYY-MM-DD format.")
            continue
        slot = slot_map.get(str(row.get("schedule") or "Morning").strip().lower(), Session.ScheduleSlot.MORNING)
        default_start_clock = time(6, 30) if slot == Session.ScheduleSlot.MORNING else time(16, 0)
        default_end_clock = time(8, 30) if slot == Session.ScheduleSlot.MORNING else time(18, 0)
        try:
            start_clock = parse_time_cell(row.get("from_time"), default_start_clock)
            end_clock = parse_time_cell(row.get("to_time"), default_end_clock)
        except ValueError as exc:
            errors.append(f"Row {index}: {exc}")
            continue
        if end_date < start_date:
            errors.append(f"Row {index}: End Date cannot be before Start Date.")
            continue
        if start_date == end_date and end_clock <= start_clock:
            errors.append(f"Row {index}: To Time must be after From Time.")
            continue
        venue = str(row.get("venue") or "").strip()
        other_venue = str(row.get("other_venue") or "").strip()
        if venue.lower() == "other" and other_venue:
            venue = f"Other - {other_venue}"
        full_day = truthy_cell(row.get("full_day"), False)
        session = Session.objects.create(
            team=team,
            title=str(row.get("title") or "Practice session").strip() or "Practice session",
            start_at=timezone.make_aware(datetime.combine(start_date, start_clock)),
            end_at=timezone.make_aware(datetime.combine(end_date, end_clock)),
            schedule_slot=slot,
            full_day=full_day,
            venue=venue,
            notes=str(row.get("notes") or "").strip(),
            scheduled_by=user,
        )
        create_notifications(
            users_for_team(team),
            "New practice session",
            f"{session.title or 'Practice session'} scheduled for {team}.",
            actor=user,
            target_url=reverse("sessions"),
            sport=team.sport,
            team=team,
            session=session,
            email_type="practice_scheduled",
            email_context={"details": [("Sport", team.sport.name), ("Team", f"{team.get_gender_display()} {team.name}"), ("Date", session.date_range_display), ("Time / Schedule", session_time_schedule_label(session)), ("Venue", session.venue), ("Scheduled By", user_label(user))]},
        )
        created += 1
    return created, errors


@login_required
@role_required(*ADMIN_ROLES)
def sports_list(request):
    if request.method == "POST":
        action = request.POST.get("action")
        sport_id = request.POST.get("sport_id")
        sport = get_object_or_404(Sport, pk=sport_id) if sport_id else None

        if action in {"create", "update"}:
            form = SportForm(request.POST, instance=sport)
            if form.is_valid():
                saved_sport = form.save()
                if action == "create":
                    notify_common_action(
                        request.user,
                        "New sport added",
                        f"{saved_sport.name} has been added to the sports department.",
                        reverse("sports"),
                        email_type="sport_created",
                        email_context={"details": [("Sport", saved_sport.name), ("Added By", user_label(request.user))]},
                    )
                messages.success(request, "Sport saved successfully.")
            else:
                messages.error(request, "Please correct the sport details and try again.")
        elif action == "bulk_upload":
            try:
                created, errors = import_sports_from_file(request.FILES.get("bulk_file"))
                add_bulk_upload_messages(request, created, errors)
            except ValueError as exc:
                messages.error(request, str(exc))
        elif action == "deactivate" and sport:
            sport.is_active = False
            sport.save(update_fields=["is_active", "updated_at"])
            messages.success(request, f"{sport.name} deactivated.")
        elif action == "activate" and sport:
            sport.is_active = True
            sport.save(update_fields=["is_active", "updated_at"])
            messages.success(request, f"{sport.name} activated.")
        elif action == "delete" and sport:
            sport_name = sport.name
            sport.delete()
            messages.success(request, f"{sport_name} deleted.")
        return redirect("sports")

    sports = Sport.objects.annotate(
        team_count=Count("teams", distinct=True),
        player_count=Count(
            "teams__memberships__user",
            filter=Q(
                teams__memberships__is_active=True,
                teams__memberships__user__profile__role=UserProfile.Role.MEMBER,
            ),
            distinct=True,
        ),
        marked_session_count=Count(
            "teams__sessions",
            filter=Q(teams__sessions__attendance_submitted=True),
            distinct=True,
        ),
    ).order_by("name")
    for sport in sports:
        sport.icon_class = sport_icon_class(sport.name)
    trainer_count = User.objects.filter(
        profile__role__in=[UserProfile.Role.TRAINER, UserProfile.Role.COORDINATOR],
        is_active=True,
    ).count()
    return render(request, "core/sports.html", {
        "sports": sports,
        "sport_count": sports.count(),
        "trainer_count": trainer_count,
    })


@login_required
@role_required(*ADMIN_ROLES)
def venues_list(request):
    if request.method == "POST":
        action = request.POST.get("action")
        venue_id = request.POST.get("venue_id")
        venue = get_object_or_404(Venue, pk=venue_id) if venue_id else None
        if action in {"create", "update"}:
            form = VenueForm(request.POST, instance=venue)
            if form.is_valid():
                saved_venue = form.save()
                if action == "create":
                    notify_common_action(
                        request.user,
                        "New venue added",
                        f"{saved_venue.name} is now available for practice sessions.",
                        reverse("venues"),
                    )
                messages.success(request, "Venue saved successfully.")
            else:
                messages.error(request, "Please correct the venue details and try again.")
        elif action == "bulk_upload":
            try:
                created, errors = import_venues_from_file(request.FILES.get("bulk_file"))
                add_bulk_upload_messages(request, created, errors)
            except ValueError as exc:
                messages.error(request, str(exc))
        elif action == "deactivate" and venue:
            venue.is_active = False
            venue.save(update_fields=["is_active", "updated_at"])
            messages.success(request, f"{venue.name} deactivated.")
        elif action == "activate" and venue:
            venue.is_active = True
            venue.save(update_fields=["is_active", "updated_at"])
            messages.success(request, f"{venue.name} activated.")
        elif action == "delete" and venue:
            venue_name = venue.name
            venue.delete()
            messages.success(request, f"{venue_name} deleted.")
        return redirect("venues")

    venues = list(Venue.objects.all().order_by("name"))
    for venue in venues:
        venue.session_count = Session.objects.filter(venue=venue.name).count()
        venue.completed_session_count = Session.objects.filter(venue=venue.name, attendance_submitted=True).count()
    return render(request, "core/venues.html", {
        "venues": venues,
        "venue_count": len(venues),
        "active_venue_count": sum(1 for venue in venues if venue.is_active),
    })


@login_required
@role_required(*ADMIN_ROLES)
def schools_list(request):
    if request.method == "POST":
        action = request.POST.get("action")
        school_id = request.POST.get("school_id")
        school = get_object_or_404(School, pk=school_id) if school_id else None
        if action in {"create", "update"}:
            form = SchoolForm(request.POST, instance=school)
            if form.is_valid():
                form.save()
                messages.success(request, "School saved successfully.")
            else:
                messages.error(request, "Please correct the school details and try again.")
        elif action == "bulk_upload":
            try:
                created, errors = import_schools_from_file(request.FILES.get("bulk_file"))
                add_bulk_upload_messages(request, created, errors)
            except ValueError as exc:
                messages.error(request, str(exc))
        elif action == "deactivate" and school:
            school.is_active = False
            school.save(update_fields=["is_active", "updated_at"])
            messages.success(request, f"{school.name} deactivated.")
        elif action == "activate" and school:
            school.is_active = True
            school.save(update_fields=["is_active", "updated_at"])
            messages.success(request, f"{school.name} activated.")
        elif action == "delete" and school:
            school_name = school.name
            school.delete()
            messages.success(request, f"{school_name} deleted.")
        return redirect("schools")

    schools = list(School.objects.annotate(student_count=Count("profiles")).order_by("name"))
    return render(request, "core/schools.html", {
        "schools": schools,
        "school_count": len(schools),
        "active_school_count": sum(1 for school in schools if school.is_active),
    })


@login_required
@role_required(*ADMIN_ROLES)
def sport_form(request, pk=None):
    sport = get_object_or_404(Sport, pk=pk) if pk else None
    form = SportForm(request.POST or None, instance=sport)
    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, "Sport saved.")
        return redirect("sports")
    return render(request, "core/form.html", {"form": form, "title": "Sport", "back_url": reverse("sports")})


@login_required
def teams_list(request):
    if is_admin_user(request.user) and request.method == "POST":
        action = request.POST.get("action")
        team_id = request.POST.get("team_id")
        team = get_object_or_404(Team, pk=team_id) if team_id else None
        if action in {"create", "update"}:
            previous_roles = {
                "captain": team.captain if team else None,
                "vice_captain": team.vice_captain if team else None,
                "coordinator": team.coordinator if team else None,
            }
            form = TeamForm(request.POST, instance=team)
            if form.is_valid():
                saved_team = form.save()
                for leader in (saved_team.captain, saved_team.vice_captain):
                    if leader:
                        membership, _ = Membership.objects.get_or_create(user=leader, team=saved_team)
                        if not membership.is_active:
                            membership.is_active = True
                            membership.save(update_fields=["is_active", "updated_at"])
                current_roles = {
                    "captain": saved_team.captain,
                    "vice_captain": saved_team.vice_captain,
                    "coordinator": saved_team.coordinator,
                }
                role_labels = {
                    "captain": "Captain",
                    "vice_captain": "Vice Captain",
                    "coordinator": "Trainer",
                }
                for role_key, role_label in role_labels.items():
                    previous_user = previous_roles[role_key]
                    current_user = current_roles[role_key]
                    if previous_user and previous_user != current_user:
                        notify_team_role_change(request.user, previous_user, saved_team, role_label, assigned=False)
                    if current_user and current_user != previous_user:
                        notify_team_role_change(request.user, current_user, saved_team, role_label, assigned=True)
                if action == "create":
                    create_notifications(
                        users_for_team(saved_team),
                        "New team added",
                        f"{saved_team} has been created.",
                        actor=request.user,
                        target_url=reverse("teams"),
                        sport=saved_team.sport,
                        team=saved_team,
                        email_type="team_created",
                        email_context={"details": [("Sport", saved_team.sport.name), ("Team", f"{saved_team.get_gender_display()} {saved_team.name}"), ("Team Type", saved_team.get_team_type_display()), ("Created By", user_label(request.user))]},
                    )
                messages.success(request, "Team saved successfully.")
            else:
                for field_errors in form.errors.values():
                    for error in field_errors:
                        messages.error(request, error)
                messages.error(request, "Please correct the team details and try again.")
        elif action == "bulk_upload":
            try:
                created, errors = import_teams_from_file(request.FILES.get("bulk_file"))
                add_bulk_upload_messages(request, created, errors)
            except ValueError as exc:
                messages.error(request, str(exc))
        elif action == "deactivate" and team:
            team.is_active = False
            team.save(update_fields=["is_active", "updated_at"])
            messages.success(request, f"{team.name} deactivated.")
        elif action == "activate" and team:
            team.is_active = True
            team.save(update_fields=["is_active", "updated_at"])
            messages.success(request, f"{team.name} activated.")
        elif action == "delete" and team:
            team_name = team.name
            team.delete()
            messages.success(request, f"{team_name} deleted.")
        elif action == "update_team_players" and team:
            selected_user_ids = {user_id for user_id in request.POST.getlist("players") if str(user_id).isdigit()}
            submitted_candidate_ids = {user_id for user_id in request.POST.getlist("all_player_ids") if str(user_id).isdigit()}
            student_roles = [UserProfile.Role.MEMBER, UserProfile.Role.CAPTAIN, UserProfile.Role.VICE_CAPTAIN]
            eligible_students = User.objects.filter(
                is_active=True,
                profile__role__in=student_roles,
            ).distinct()
            eligible_ids = {str(user_id) for user_id in eligible_students.values_list("pk", flat=True)}
            submitted_candidate_ids &= eligible_ids
            if not submitted_candidate_ids:
                messages.error(request, "No eligible students were submitted for this team update.")
                return redirect("teams")
            selected_user_ids &= eligible_ids
            selected_user_ids &= submitted_candidate_ids
            conflict_messages = []
            for user_id in selected_user_ids:
                student = User.objects.get(pk=user_id)
                conflict = validate_student_team_memberships(student, [team], replace_team_ids=[team.pk])
                if conflict:
                    student_name = student.get_full_name() or student.email or student.username
                    conflict_messages.append(f"{student_name}: {conflict}")
            if conflict_messages:
                messages.error(request, " ".join(conflict_messages[:3]))
                return redirect("teams")
            added_ids = set()
            activated_ids = set()
            removed_ids = set()
            with transaction.atomic():
                existing_memberships = {
                    item.user_id: item.is_active
                    for item in Membership.objects.filter(team=team)
                }
                existing_ids = set(existing_memberships)
                previous_active_ids = {user_id for user_id, is_active in existing_memberships.items() if is_active}
                selected_ids_int = {int(user_id) for user_id in selected_user_ids}
                submitted_ids_int = {int(user_id) for user_id in submitted_candidate_ids}
                remove_ids = submitted_ids_int - selected_ids_int
                added_ids = selected_ids_int - existing_ids
                activated_ids = {user_id for user_id in selected_ids_int.intersection(existing_ids) if not existing_memberships.get(user_id)}
                removed_ids = previous_active_ids.intersection(remove_ids)
                Membership.objects.filter(team=team, user_id__in=remove_ids).delete()
                for user_id in selected_user_ids:
                    membership, _ = Membership.objects.get_or_create(user_id=user_id, team=team)
                    if not membership.is_active:
                        membership.is_active = True
                        membership.save(update_fields=["is_active", "updated_at"])
                updates = []
                if team.captain_id and str(team.captain_id) not in selected_user_ids:
                    team.captain = None
                    updates.append("captain")
                if team.vice_captain_id and str(team.vice_captain_id) not in selected_user_ids:
                    team.vice_captain = None
                    updates.append("vice_captain")
                if updates:
                    updates.append("updated_at")
                    team.save(update_fields=updates)
                added_count = len(added_ids) + len(activated_ids)
                removed_count = len(removed_ids)
            notify_team_membership_change(request.user, User.objects.filter(pk__in=added_ids), team, "added")
            notify_team_membership_change(request.user, User.objects.filter(pk__in=activated_ids), team, "activated")
            notify_team_membership_change(request.user, User.objects.filter(pk__in=removed_ids), team, "removed")
            messages.success(request, f"Team players updated successfully. Added {added_count}, removed {removed_count}.")
        return redirect("teams")

    teams = Team.objects.select_related("sport", "captain", "vice_captain", "coordinator").prefetch_related("memberships__user", "memberships__user__profile")
    if not is_admin_user(request.user):
        teams = teams.filter(Q(memberships__user=request.user) | Q(captain=request.user) | Q(vice_captain=request.user) | Q(coordinator=request.user)).distinct()
    teams = list(teams)
    visible_team_ids = [team.pk for team in teams]
    team_stats = {
        "total": len(teams),
        "completed_sessions": Session.objects.filter(team_id__in=visible_team_ids, attendance_submitted=True).count(),
    }
    student_roles = [UserProfile.Role.MEMBER, UserProfile.Role.CAPTAIN, UserProfile.Role.VICE_CAPTAIN]
    can_manage_team_players = is_admin_user(request.user)
    all_student_candidates = list(User.objects.filter(
        is_active=True,
        profile__role__in=student_roles,
    ).distinct().select_related("profile").prefetch_related("memberships__team").order_by("first_name", "last_name", "email"))
    for team in teams:
        team.active_memberships = [membership for membership in team.memberships.all() if membership.is_active]
        team.active_member_ids = {membership.user_id for membership in team.active_memberships}
        team.player_count = len(team.active_member_ids)
        team.scheduled_session_count = team.sessions.count()
        team.completed_session_count = team.sessions.filter(attendance_submitted=True).count()
        team.can_manage_players = can_manage_team_players
        team.can_view_players = (
            can_manage_team_players
            or request.user == team.coordinator
            or request.user == team.captain
            or request.user == team.vice_captain
        )
        if can_manage_team_players:
            team.player_candidates = sorted(
                all_student_candidates,
                key=lambda student: (
                    student.pk not in team.active_member_ids,
                    (student.get_full_name() or student.email or student.username).lower(),
                ),
            )
        else:
            team.player_candidates = sorted(
                [membership.user for membership in team.active_memberships],
                key=lambda student: (student.get_full_name() or student.email or student.username).lower(),
            )
    trainer_roles = [UserProfile.Role.TRAINER, UserProfile.Role.COORDINATOR]
    users = User.objects.filter(is_active=True, profile__role__in=student_roles).prefetch_related("memberships__team").order_by("first_name", "last_name", "username")
    for student in users:
        sport_ids = {
            str(membership.team.sport_id)
            for membership in student.memberships.all()
            if membership.is_active
        }
        student.assigned_sport_ids_csv = ",".join(sorted(sport_ids))
    trainer_users = User.objects.filter(is_active=True, profile__role__in=trainer_roles).order_by("first_name", "last_name", "username")
    sports = Sport.objects.filter(is_active=True).order_by("name")
    return render(request, "core/teams.html", {
        "teams": teams,
        "users": users,
        "trainer_users": trainer_users,
        "sports": sports,
        "team_stats": team_stats,
    })


@login_required
@role_required(*ADMIN_ROLES)
def team_form(request, pk=None):
    team = get_object_or_404(Team, pk=pk) if pk else None
    form = TeamForm(request.POST or None, instance=team)
    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, "Team saved.")
        return redirect("teams")
    return render(request, "core/form.html", {"form": form, "title": "Team", "back_url": reverse("teams")})


@login_required
def members_list(request):
    can_manage_students = is_admin_user(request.user) or role_for(request.user) in TRAINER_ROLES
    allowed_teams = trainer_assigned_teams(request.user)
    if can_manage_students and request.method == "POST":
        action = request.POST.get("action")
        request_id = request.POST.get("request_id")
        access_request = get_object_or_404(LoginAccessRequest, pk=request_id) if request_id else None
        if action in {"approve_request", "reject_request"} and access_request:
            if not is_admin_user(request.user):
                messages.error(request, "You cannot manage login requests.")
                return redirect("members")
            if action == "approve_request":
                try:
                    validate_christ_email(access_request.email)
                except ValueError as exc:
                    messages.error(request, str(exc))
                    return redirect("members")
            with transaction.atomic():
                access_request.status = LoginAccessRequest.Status.APPROVED if action == "approve_request" else LoginAccessRequest.Status.REJECTED
                access_request.reviewed_by = request.user
                access_request.reviewed_at = timezone.now()
                access_request.save(update_fields=["status", "reviewed_by", "reviewed_at", "updated_at"])
                if action == "approve_request":
                    create_student_from_access_request(access_request)
            message = "approved and added to Add Students" if action == "approve_request" else "rejected"
            messages.success(request, f"{access_request.email} {message}.")
            return redirect("members")

        if action == "bulk_upload":
            try:
                created, errors = import_students_from_file(request.FILES.get("bulk_file"), request.user)
                add_bulk_upload_messages(request, created, errors)
            except ValueError as exc:
                messages.error(request, str(exc))
            return redirect("members")

        membership_id = request.POST.get("membership_id")
        membership = get_object_or_404(Membership, pk=membership_id) if membership_id else None
        status_reason = (request.POST.get("status_reason") or "").strip()
        if membership and not can_manage_student_membership(request.user, membership):
            messages.error(request, "You cannot manage students from this team.")
            return redirect("members")
        user_id = request.POST.get("user_id")
        existing_user = membership.user if membership else User.objects.filter(pk=user_id).first() if user_id else None
        if existing_user and not is_admin_user(request.user) and not can_view_student_user(request.user, existing_user):
            messages.error(request, "You cannot manage this student.")
            return redirect("members")
        if action in {"create", "update"}:
            membership_notifications = []
            try:
                with transaction.atomic():
                    posted_email = request.POST.get("student_email", "").strip().lower()
                    email_user = User.objects.filter(email__iexact=posted_email).first() if posted_email else None
                    if not is_admin_user(request.user) and email_user and not can_view_student_user(request.user, email_user):
                        if action == "update":
                            messages.error(request, "You cannot update students from another team.")
                            return redirect("members")
                        student = email_user
                    else:
                        student = build_student_from_post(request, existing_user=existing_user)
                    previous_active_team_ids = set(Membership.objects.filter(
                        user=student,
                        team__in=allowed_teams,
                        is_active=True,
                    ).values_list("team_id", flat=True))
                    selected_sport_ids = {int(sport_id) for sport_id in request.POST.getlist("sports") if str(sport_id).isdigit()}
                    selected_team_types = set(request.POST.getlist("team_types"))
                    valid_team_types = {value for value, _ in Team.TeamType.choices}
                    selected_team_types = selected_team_types.intersection(valid_team_types)
                    selected_team_ids = request.POST.getlist("teams")
                    fallback_team = request.POST.get("team")
                    if not selected_team_ids and fallback_team:
                        selected_team_ids = [fallback_team]
                    allowed_team_ids = set(allowed_teams.values_list("pk", flat=True))
                    selected_team_ids = {int(team_id) for team_id in selected_team_ids if str(team_id).isdigit()}
                    if selected_sport_ids:
                        selected_team_ids = set(
                            Team.objects.filter(pk__in=selected_team_ids, sport_id__in=selected_sport_ids).values_list("pk", flat=True)
                        )
                    if selected_team_types:
                        selected_team_ids = set(
                            Team.objects.filter(pk__in=selected_team_ids, team_type__in=selected_team_types).values_list("pk", flat=True)
                        )
                    if selected_team_ids and not selected_team_ids.issubset(allowed_team_ids):
                        messages.error(request, "You cannot assign students to one or more selected teams.")
                        return redirect("members")
                    if selected_team_ids:
                        selected_teams = Team.objects.filter(pk__in=selected_team_ids).select_related("sport")
                        conflict = validate_student_team_memberships(student, selected_teams, replace_team_ids=allowed_team_ids)
                        if conflict:
                            messages.error(request, conflict)
                            return redirect("members")
                        active_value = request.POST.get("is_active") == "on"
                        for selected_team in selected_teams:
                            selected_membership, _ = Membership.objects.get_or_create(user=student, team=selected_team)
                            selected_membership.is_active = active_value
                            selected_membership.save()
                        Membership.objects.filter(user=student, team_id__in=allowed_team_ids).exclude(team_id__in=selected_team_ids).delete()
                        messages.success(request, "Student saved and team assignments updated successfully.")
                    else:
                        if not is_admin_user(request.user):
                            messages.error(request, "Team is required for trainer student access.")
                            return redirect("members")
                        if action == "update":
                            Membership.objects.filter(user=student, team_id__in=allowed_team_ids).delete()
                        messages.success(request, "Student saved. Team can be assigned later.")
                    current_memberships = {
                        item.team_id: item.is_active
                        for item in Membership.objects.filter(user=student, team__in=allowed_teams)
                    }
                    current_active_team_ids = {team_id for team_id, is_active in current_memberships.items() if is_active}
                    added_or_activated_ids = current_active_team_ids - previous_active_team_ids
                    no_longer_active_ids = previous_active_team_ids - current_active_team_ids
                    deactivated_ids = {team_id for team_id in no_longer_active_ids if team_id in current_memberships}
                    removed_ids = no_longer_active_ids - deactivated_ids
                    changed_team_ids = added_or_activated_ids | deactivated_ids | removed_ids
                    teams_by_id = {
                        item.pk: item
                        for item in Team.objects.filter(pk__in=changed_team_ids).select_related("sport")
                    }
                    membership_notifications = [
                        (teams_by_id[team_id], "added")
                        for team_id in added_or_activated_ids
                        if team_id in teams_by_id
                    ]
                    membership_notifications.extend(
                        (teams_by_id[team_id], "deactivated")
                        for team_id in deactivated_ids
                        if team_id in teams_by_id
                    )
                    membership_notifications.extend(
                        (teams_by_id[team_id], "removed")
                        for team_id in removed_ids
                        if team_id in teams_by_id
                    )
            except ValueError as exc:
                messages.error(request, str(exc))
            else:
                for changed_team, change in membership_notifications:
                    notify_team_membership_change(request.user, [student], changed_team, change)
        elif action == "deactivate" and membership:
            membership.is_active = False
            membership.save(update_fields=["is_active", "updated_at"])
            notify_team_membership_change(request.user, [membership.user], membership.team, "deactivated")
            messages.success(request, "Member deactivated for this team.")
        elif action == "deactivate" and existing_user and role_for(request.user) in TRAINER_ROLES:
            memberships_to_deactivate = list(Membership.objects.filter(user=existing_user, team__in=allowed_teams, is_active=True).select_related("team", "team__sport"))
            Membership.objects.filter(user=existing_user, team__in=allowed_teams).update(is_active=False)
            set_user_account_status(existing_user, UserProfile.AccountStatus.DEACTIVATED, status_reason)
            for changed_membership in memberships_to_deactivate:
                notify_team_membership_change(request.user, [existing_user], changed_membership.team, "deactivated")
            messages.success(request, "Student deactivated for your assigned team(s).")
        elif action == "deactivate" and existing_user and is_admin_user(request.user):
            set_user_account_status(existing_user, UserProfile.AccountStatus.DEACTIVATED, status_reason)
            messages.success(request, "Student deactivated.")
        elif action == "activate" and membership:
            membership.is_active = True
            membership.save(update_fields=["is_active", "updated_at"])
            notify_team_membership_change(request.user, [membership.user], membership.team, "activated")
            messages.success(request, "Member activated for this team.")
        elif action == "activate" and existing_user and role_for(request.user) in TRAINER_ROLES:
            memberships_to_activate = list(Membership.objects.filter(user=existing_user, team__in=allowed_teams, is_active=False).select_related("team", "team__sport"))
            Membership.objects.filter(user=existing_user, team__in=allowed_teams).update(is_active=True)
            set_user_account_status(existing_user, UserProfile.AccountStatus.ACTIVE)
            for changed_membership in memberships_to_activate:
                notify_team_membership_change(request.user, [existing_user], changed_membership.team, "activated")
            messages.success(request, "Student activated for your assigned team(s).")
        elif action == "activate" and existing_user and is_admin_user(request.user):
            set_user_account_status(existing_user, UserProfile.AccountStatus.ACTIVE)
            messages.success(request, "Student activated.")
        elif action == "delete" and membership:
            changed_team = membership.team
            changed_user = membership.user
            membership.delete()
            notify_team_membership_change(request.user, [changed_user], changed_team, "removed")
            messages.success(request, "Member assignment deleted.")
        elif action == "delete" and existing_user and role_for(request.user) in TRAINER_ROLES:
            memberships_to_delete = list(Membership.objects.filter(user=existing_user, team__in=allowed_teams).select_related("team", "team__sport"))
            set_user_account_status(existing_user, UserProfile.AccountStatus.DELETED, status_reason)
            for changed_membership in memberships_to_delete:
                notify_team_membership_change(request.user, [existing_user], changed_membership.team, "removed")
            messages.success(request, "Student removed from your assigned team(s).")
        elif action == "delete" and existing_user and is_admin_user(request.user):
            student_name = existing_user.get_full_name() or existing_user.email or existing_user.username
            set_user_account_status(existing_user, UserProfile.AccountStatus.DELETED, status_reason)
            messages.success(request, f"{student_name} moved to settings restore list.")
        return redirect("members")

    memberships = Membership.objects.select_related("user", "user__profile", "team", "team__sport").exclude(user__profile__account_status=UserProfile.AccountStatus.DELETED)
    if not is_admin_user(request.user):
        if role_for(request.user) in TRAINER_ROLES:
            memberships = memberships.filter(team__in=allowed_teams)
        else:
            memberships = memberships.filter(team__in=Team.objects.filter(Q(captain=request.user) | Q(vice_captain=request.user)))
    assigned_user_ids = set()
    assigned_team_ids_by_user = {}
    assigned_team_labels_by_user = {}
    assigned_sport_ids_by_user = {}
    assigned_sport_labels_by_user = {}
    assigned_team_types_by_user = {}
    assigned_team_type_labels_by_user = {}
    active_membership_by_user = {}
    first_membership_by_user = {}
    team_type_labels = dict(Team.TeamType.choices)
    for assigned_membership in memberships:
        first_membership_by_user.setdefault(assigned_membership.user_id, assigned_membership)
        active_membership_by_user[assigned_membership.user_id] = (
            active_membership_by_user.get(assigned_membership.user_id, False) or assigned_membership.is_active
        )
        assigned_team_ids_by_user.setdefault(assigned_membership.user_id, []).append(str(assigned_membership.team_id))
        assigned_team_labels_by_user.setdefault(assigned_membership.user_id, []).append(
            f"{assigned_membership.team.sport.name} - {assigned_membership.team.get_gender_display()} {assigned_membership.team.name}"
        )
        assigned_sport_ids_by_user.setdefault(assigned_membership.user_id, set()).add(str(assigned_membership.team.sport_id))
        assigned_sport_labels_by_user.setdefault(assigned_membership.user_id, set()).add(assigned_membership.team.sport.name)
        assigned_team_types_by_user.setdefault(assigned_membership.user_id, set()).add(assigned_membership.team.team_type)
        assigned_team_type_labels_by_user.setdefault(assigned_membership.user_id, set()).add(team_type_labels.get(assigned_membership.team.team_type, assigned_membership.team.team_type))
    student_rows = []
    for user_id, membership in first_membership_by_user.items():
        assigned_user_ids.add(user_id)
        account_status = user_account_status(membership.user)
        student_rows.append({
            "membership": None,
            "user": membership.user,
            "team": None,
            "account_status": account_status,
            "is_readonly": account_status == UserProfile.AccountStatus.DEACTIVATED,
            "is_active": account_status == UserProfile.AccountStatus.ACTIVE and active_membership_by_user.get(user_id, False),
            "assigned_team_ids": ",".join(assigned_team_ids_by_user.get(user_id, [])),
            "assigned_team_labels": "; ".join(assigned_team_labels_by_user.get(user_id, [])),
            "assigned_sport_ids": ",".join(sorted(assigned_sport_ids_by_user.get(user_id, set()))),
            "assigned_sport_labels": ", ".join(sorted(assigned_sport_labels_by_user.get(user_id, set()))),
            "assigned_team_types": ",".join(sorted(assigned_team_types_by_user.get(user_id, set()))),
            "assigned_team_type_labels": ", ".join(sorted(assigned_team_type_labels_by_user.get(user_id, set()))),
        })
    if is_admin_user(request.user):
        unassigned_students = User.objects.filter(profile__role=UserProfile.Role.MEMBER).exclude(profile__account_status=UserProfile.AccountStatus.DELETED).exclude(pk__in=assigned_user_ids).select_related("profile").order_by("first_name", "last_name", "email")
        for student in unassigned_students:
            account_status = user_account_status(student)
            student_rows.append({"membership": None, "user": student, "team": None, "account_status": account_status, "is_readonly": account_status == UserProfile.AccountStatus.DEACTIVATED, "is_active": account_status == UserProfile.AccountStatus.ACTIVE, "assigned_team_ids": "", "assigned_team_labels": "", "assigned_sport_ids": "", "assigned_sport_labels": "", "assigned_team_types": "", "assigned_team_type_labels": ""})
    student_rows.sort(key=lambda row: (row["account_status"] == UserProfile.AccountStatus.DEACTIVATED, (row["user"].get_full_name() or row["user"].email or row["user"].username).lower()))
    teams = allowed_teams.select_related("sport").order_by("sport__name", "name")
    sports = Sport.objects.filter(teams__in=allowed_teams, is_active=True).distinct().order_by("name")
    schools = School.objects.filter(is_active=True).order_by("name")
    student_stats = {
        "total": len(student_rows),
        "active": sum(1 for row in student_rows if row["account_status"] == UserProfile.AccountStatus.ACTIVE and row["is_active"]),
    }
    login_requests = LoginAccessRequest.objects.exclude(status=LoginAccessRequest.Status.APPROVED).order_by("status", "-requested_at") if is_admin_user(request.user) else []
    registration_numbers = list(UserProfile.objects.exclude(register_no="").values("user_id", "register_no"))
    return render(request, "core/members.html", {
        "memberships": memberships,
        "student_rows": student_rows,
        "teams": teams,
        "sports": sports,
        "schools": schools,
        "team_types": Team.TeamType.choices,
        "login_requests": login_requests,
        "can_manage_students": can_manage_students,
        "is_admin_user": is_admin_user(request.user),
        "registration_numbers": registration_numbers,
        "student_stats": student_stats,
    })


@login_required
def member_detail(request, pk):
    student = get_object_or_404(User.objects.select_related("profile"), pk=pk)
    if not can_view_student_user(request.user, student):
        messages.error(request, "You cannot view this student.")
        return redirect("members")
    memberships = Membership.objects.filter(user=student).select_related("team", "team__sport")
    records = AttendanceRecord.objects.filter(member=student).select_related("session", "session__team", "session__team__sport", "marked_by")
    if not is_admin_user(request.user):
        allowed_teams = trainer_assigned_teams(request.user)
        memberships = memberships.filter(team__in=allowed_teams)
        records = records.filter(session__team__in=allowed_teams)
    total_records = records.count()
    attended_records = records.filter(status__in=[
        AttendanceRecord.Status.PRESENT,
        AttendanceRecord.Status.LATE,
        AttendanceRecord.Status.EARLY_EXIT,
    ]).count()
    attendance_percent = round((attended_records / total_records) * 100, 1) if total_records else None
    attendance_sport_stats = []
    sport_ids = set(memberships.values_list("team__sport_id", flat=True))
    for sport in Sport.objects.filter(pk__in=sport_ids).order_by("name"):
        sport_records = records.filter(session__team__sport=sport)
        sport_total = sport_records.count()
        sport_attended = sport_records.filter(status__in=[
            AttendanceRecord.Status.PRESENT,
            AttendanceRecord.Status.LATE,
            AttendanceRecord.Status.EARLY_EXIT,
        ]).count()
        attendance_sport_stats.append({
            "sport": sport,
            "total": sport_total,
            "attended": sport_attended,
            "percent": round((sport_attended / sport_total) * 100, 1) if sport_total else None,
        })
    breadcrumb_items = [
        {"label": "Dashboard", "url_name": "dashboard"},
        {"label": "Add Students", "url_name": "members"},
        {"label": student.get_full_name() or student.email or student.username},
    ]
    return render(request, "core/member_detail.html", {
        "student": student,
        "memberships": memberships,
        "records": records,
        "attendance_percent": attendance_percent,
        "attendance_total": total_records,
        "attendance_attended": attended_records,
        "attendance_sport_stats": attendance_sport_stats,
        "breadcrumb_items": breadcrumb_items,
    })


@login_required
@role_required(*ADMIN_ROLES)
def trainers_list(request):
    if request.method == "POST":
        action = request.POST.get("action")
        trainer_id = request.POST.get("trainer_id")
        trainer = get_object_or_404(User, pk=trainer_id) if trainer_id else None
        status_reason = (request.POST.get("status_reason") or "").strip()
        if action in {"create", "update"}:
            full_name = request.POST.get("trainer_name", "").strip()
            email = request.POST.get("trainer_email", "").strip().lower()
            phone = request.POST.get("mobile_number", "").strip()
            register_no = request.POST.get("register_no", "").strip()
            selected_team_ids = request.POST.getlist("teams")
            is_new_trainer = trainer is None and not User.objects.filter(email__iexact=email).exists()
            if not email:
                messages.error(request, "Trainer email is required.")
                return redirect("trainers")
            try:
                email = validate_christ_email(email)
            except ValueError as exc:
                messages.error(request, str(exc))
                return redirect("trainers")
            if phone and not re.fullmatch(r"\d{10}", phone):
                messages.error(request, "Mobile number must be exactly 10 digits.")
                return redirect("trainers")
            trainer = trainer or User.objects.filter(email__iexact=email).first()
            previous_team_ids = set(Team.objects.filter(coordinator=trainer).values_list("pk", flat=True)) if trainer else set()
            try:
                validate_unique_register_no(register_no, trainer)
            except ValueError as exc:
                messages.error(request, str(exc))
                return redirect("trainers")
            if trainer is None:
                trainer = User(username=email, email=email)
                trainer.set_unusable_password()
            if full_name:
                parts = full_name.split(" ", 1)
                trainer.first_name = parts[0]
                trainer.last_name = parts[1] if len(parts) > 1 else ""
            trainer.email = email
            trainer.username = email
            trainer.save()
            profile, _ = UserProfile.objects.get_or_create(user=trainer)
            profile.role = UserProfile.Role.TRAINER
            profile.phone = phone
            profile.register_no = register_no
            profile.save(update_fields=["role", "phone", "register_no", "updated_at"])
            Team.objects.filter(coordinator=trainer).exclude(pk__in=selected_team_ids).update(coordinator=None)
            Team.objects.filter(pk__in=selected_team_ids).update(coordinator=trainer)
            selected_ids = set(Team.objects.filter(pk__in=selected_team_ids).values_list("pk", flat=True))
            changed_teams = Team.objects.filter(pk__in=previous_team_ids | selected_ids).select_related("sport")
            for changed_team in changed_teams:
                if changed_team.pk in previous_team_ids and changed_team.pk not in selected_ids:
                    notify_team_role_change(request.user, trainer, changed_team, "Trainer", assigned=False)
                elif changed_team.pk in selected_ids and changed_team.pk not in previous_team_ids:
                    notify_team_role_change(request.user, trainer, changed_team, "Trainer", assigned=True)
            if action == "create" and is_new_trainer:
                notify_common_action(
                    request.user,
                    "New trainer added",
                    f"{trainer.get_full_name() or trainer.email} has been added as a trainer.",
                    reverse("trainers"),
                )
            messages.success(request, "Trainer saved and assigned successfully.")
        elif action == "bulk_upload":
            try:
                created, errors = import_trainers_from_file(request.FILES.get("bulk_file"))
                add_bulk_upload_messages(request, created, errors)
            except ValueError as exc:
                messages.error(request, str(exc))
        elif action == "deactivate" and trainer:
            set_user_account_status(trainer, UserProfile.AccountStatus.DEACTIVATED, status_reason)
            messages.success(request, "Trainer deactivated.")
        elif action == "activate" and trainer:
            set_user_account_status(trainer, UserProfile.AccountStatus.ACTIVE)
            messages.success(request, "Trainer activated.")
        elif action == "delete" and trainer:
            trainer_name = trainer.get_full_name() or trainer.username
            set_user_account_status(trainer, UserProfile.AccountStatus.DELETED, status_reason)
            messages.success(request, f"{trainer_name} moved to settings restore list.")
        return redirect("trainers")

    trainers = User.objects.filter(
        Q(profile__role__in=[UserProfile.Role.TRAINER, UserProfile.Role.COORDINATOR])
        | Q(coordinated_teams__isnull=False)
    ).exclude(profile__account_status=UserProfile.AccountStatus.DELETED).annotate(
        assigned_session_count=Count("coordinated_teams__sessions", distinct=True),
        marked_session_count=Count("attendance_marked__session", distinct=True),
        feedback_received_count=Count("feedback_received", distinct=True),
        feedback_submitted_count=Count("feedback_sent", distinct=True),
    ).distinct().prefetch_related("coordinated_teams__sport").order_by("first_name", "last_name", "email")
    for trainer in trainers:
        trainer.account_status = user_account_status(trainer)
        trainer.is_readonly = trainer.account_status == UserProfile.AccountStatus.DEACTIVATED
        sport_names = sorted({team.sport.name for team in trainer.coordinated_teams.all()})
        trainer.assigned_sport_names = sport_names
        trainer.assigned_sports_label = ", ".join(sport_names) if sport_names else "No sports assigned"
    trainers = sorted(trainers, key=lambda trainer: (trainer.account_status == UserProfile.AccountStatus.DEACTIVATED, (trainer.get_full_name() or trainer.email or trainer.username).lower()))
    trainer_stats = {
        "total": len(trainers),
        "active": sum(1 for trainer in trainers if trainer.account_status == UserProfile.AccountStatus.ACTIVE and trainer.profile.role in [UserProfile.Role.TRAINER, UserProfile.Role.COORDINATOR]),
        "assigned_sports": Sport.objects.filter(
            teams__coordinator__is_active=True,
            teams__coordinator__profile__account_status=UserProfile.AccountStatus.ACTIVE,
        ).distinct().count(),
    }
    teams = Team.objects.filter(is_active=True).select_related("sport").order_by("sport__name", "gender", "name")
    registration_numbers = list(UserProfile.objects.exclude(register_no="").values("user_id", "register_no"))
    return render(request, "core/trainers.html", {
        "trainers": trainers,
        "teams": teams,
        "registration_numbers": registration_numbers,
        "trainer_stats": trainer_stats,
    })


@login_required
def settings_page(request):
    inactive_users = User.objects.filter(
        Q(profile__account_status__in=[UserProfile.AccountStatus.DEACTIVATED, UserProfile.AccountStatus.DELETED])
        | Q(is_active=False)
    ).select_related("profile").distinct().order_by("first_name", "last_name", "email")
    notification_count = Notification.objects.filter(user=request.user).count()
    return render(request, "core/settings.html", {
        "inactive_users_count": inactive_users.count(),
        "notification_count": notification_count,
        "is_settings_admin": is_admin_user(request.user),
    })


@login_required
@role_required(*ADMIN_ROLES)
def deactivated_users(request):
    inactive_users = User.objects.filter(
        Q(profile__account_status__in=[UserProfile.AccountStatus.DEACTIVATED, UserProfile.AccountStatus.DELETED])
        | Q(is_active=False)
    ).select_related("profile").distinct().order_by("profile__account_status", "first_name", "last_name", "email")
    for item in inactive_users:
        item.recycle_status = user_account_status(item)
    return render(request, "core/deactivated_users.html", {"inactive_users": inactive_users})


@login_required
@role_required(*ADMIN_ROLES)
def restore_user(request, pk):
    user = get_object_or_404(User, pk=pk)
    set_user_account_status(user, UserProfile.AccountStatus.ACTIVE)
    Membership.objects.filter(user=user).update(is_active=True)
    messages.success(request, f"{user.get_full_name() or user.email or user.username} restored.")
    return redirect("deactivated_users")


@login_required
@role_required(*ADMIN_ROLES)
def permanent_delete_user(request, pk):
    user = get_object_or_404(User, pk=pk)
    if request.method == "POST":
        user_name = user.get_full_name() or user.email or user.username
        user.delete()
        messages.success(request, f"{user_name} permanently deleted.")
    return redirect("deactivated_users")


@login_required
def my_profile(request):
    profile, _ = UserProfile.objects.get_or_create(user=request.user)
    if request.method == "POST" and request.POST.get("remove_profile_image") == "1":
        if profile.profile_image:
            profile.profile_image.delete(save=False)
        profile.profile_image = None
        profile.save(update_fields=["profile_image", "updated_at"])
        messages.success(request, "Profile photo removed.")
        return redirect("my_profile")

    form = ProfileForm(request.POST or None, request.FILES or None, instance=profile, user_instance=request.user)
    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, "Profile updated.")
        return redirect("my_profile")
    profile_associations = []
    if is_admin_user(request.user):
        associated_sports_count = Sport.objects.filter(is_active=True).count()
        associated_team_count = Team.objects.filter(is_active=True).count()
        profile_associations.append({
            "label": "All active sports and teams",
            "role": "Admin",
        })
    elif role_for(request.user) in TRAINER_ROLES:
        trainer_teams = Team.objects.filter(coordinator=request.user, is_active=True).select_related("sport").order_by("sport__name", "gender", "name")
        associated_sports_count = trainer_teams.values("sport").distinct().count()
        associated_team_count = trainer_teams.count()
        profile_associations = [
            {
                "label": f"{team.sport.name} - {team.get_gender_display()} {team.name}",
                "role": "Trainer",
            }
            for team in trainer_teams
        ]
    else:
        profile_memberships = Membership.objects.filter(user=request.user, is_active=True).select_related("team", "team__sport").order_by("team__sport__name", "team__gender", "team__name")
        associated_sports_count = Sport.objects.filter(teams__memberships__user=request.user, teams__memberships__is_active=True).distinct().count()
        associated_team_count = profile_memberships.count()
        for membership in profile_memberships:
            role_label = ""
            if membership.team.captain_id == request.user.pk:
                role_label = "Captain"
            elif membership.team.vice_captain_id == request.user.pk:
                role_label = "Vice Captain"
            profile_associations.append({
                "label": f"{membership.team.sport.name} - {membership.team.get_gender_display()} {membership.team.name}",
                "role": role_label,
            })
    return render(request, "core/my_profile.html", {
        "form": form,
        "profile_associations": profile_associations,
        "associated_sports_count": associated_sports_count,
        "associated_team_count": associated_team_count,
    })


@login_required
def notifications_list(request):
    notifications = Notification.objects.filter(user=request.user).select_related("actor", "sport", "team", "session")
    return render(request, "core/notifications.html", {"notifications": notifications})


@login_required
def mark_notifications_read(request):
    if request.method == "POST":
        notification_id = request.POST.get("notification_id")
        qs = Notification.objects.filter(user=request.user, is_read=False)
        if notification_id:
            qs = qs.filter(pk=notification_id)
        qs.update(is_read=True, read_at=timezone.now())
    next_url = request.POST.get("next") or request.META.get("HTTP_REFERER") or reverse("notifications")
    return redirect(next_url)


@login_required
@role_required(*ADMIN_ROLES)
def membership_form(request, pk=None):
    membership = get_object_or_404(Membership, pk=pk) if pk else None
    form = MembershipForm(request.POST or None, instance=membership)
    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, "Member assignment saved.")
        return redirect("members")
    return render(request, "core/form.html", {"form": form, "title": "Team Member", "back_url": reverse("members")})


@login_required
def sessions_list(request):
    if request.method == "POST":
        action = request.POST.get("action")
        session_id = request.POST.get("session_id")
        session = get_object_or_404(Session, pk=session_id) if session_id else None
        if session and action != "create" and not can_manage_team(request.user, session.team):
            messages.error(request, "You cannot manage this session.")
            return redirect("sessions")
        if action == "bulk_upload":
            try:
                created, errors = import_sessions_from_file(request.FILES.get("bulk_file"), request.user)
                add_bulk_upload_messages(request, created, errors)
            except (ValueError, TypeError) as exc:
                messages.error(request, str(exc))
        elif action in {"create", "update"}:
            original_team = session.team if session else None
            original_schedule = (session.start_at, session.end_at, session.schedule_slot, session.full_day, session.team_id) if session else None
            form = SessionForm(request.POST, instance=session)
            if not is_admin_user(request.user):
                form.fields["team"].queryset = Team.objects.filter(coordinator=request.user)
            if form.is_valid():
                obj = form.save(commit=False)
                if can_schedule_session(request.user, obj.team):
                    is_new_session = action == "create" or obj.pk is None
                    obj.scheduled_by = obj.scheduled_by or request.user
                    obj.save()
                    if is_new_session:
                        notify_practice_session(
                            obj,
                            "New practice session",
                            f"{obj.title or 'Practice session'} scheduled for {obj.team}.",
                            actor=request.user,
                            email_type="practice_scheduled",
                        )
                    else:
                        updated_schedule = (obj.start_at, obj.end_at, obj.schedule_slot, obj.full_day, obj.team_id)
                        event_title = "Practice session rescheduled" if original_schedule != updated_schedule else "Practice session updated"
                        impacted_teams = [obj.team]
                        if original_team and original_team.pk != obj.team_id:
                            impacted_teams.append(original_team)
                        notify_practice_session(
                            obj,
                            event_title,
                            f"{obj.title or 'Practice session'} updated for {obj.team}.",
                            actor=request.user,
                            teams=impacted_teams,
                            email_type="practice_rescheduled" if original_schedule != updated_schedule else "practice_updated",
                        )
                    messages.success(request, "Practice session saved successfully.")
                else:
                    messages.error(request, "You cannot schedule for this team.")
            else:
                for field_errors in form.errors.values():
                    for error in field_errors:
                        messages.error(request, error)
                messages.error(request, "Please correct the session details and try again.")
        elif action == "delete" and session:
            cancelled_title = session.title or "Practice session"
            cancelled_team = session.team
            session.status = Session.Status.CANCELLED
            session.attendance_started_by = None
            session.attendance_started_by_role = ""
            session.attendance_started_at = None
            session.attendance_lock_expires_at = None
            session.save(update_fields=[
                "status",
                "attendance_started_by",
                "attendance_started_by_role",
                "attendance_started_at",
                "attendance_lock_expires_at",
                "updated_at",
            ])
            notify_practice_session(
                session,
                "Practice session cancelled",
                f"{cancelled_title} cancelled for {cancelled_team}.",
                actor=request.user,
                teams=[cancelled_team],
                include_session=False,
                email_type="practice_cancelled",
            )
            messages.success(request, "Practice session cancelled.")
        return redirect("sessions")

    form = ReportFilterForm(request.GET or None)
    all_sessions = visible_sessions(request.user)
    session_stats = {
        "today": all_sessions.filter(start_at__date=timezone.localdate()).exclude(status=Session.Status.CANCELLED).count(),
        "upcoming": all_sessions.filter(start_at__date__gte=timezone.localdate(), attendance_submitted=False).exclude(status=Session.Status.CANCELLED).count(),
        "completed": all_sessions.filter(attendance_submitted=True).exclude(status=Session.Status.CANCELLED).count(),
        "cancelled": all_sessions.filter(status=Session.Status.CANCELLED).count(),
    }
    sessions = all_sessions
    if form.is_valid():
        sport = form.cleaned_data.get("sport")
        gender = form.cleaned_data.get("gender")
        team_type = form.cleaned_data.get("team_type")
        team = form.cleaned_data.get("team")
        start = form.cleaned_data.get("start_date")
        end = form.cleaned_data.get("end_date")
        if sport:
            sessions = sessions.filter(team__sport=sport)
        if gender:
            sessions = sessions.filter(team__gender=gender)
        if team_type:
            sessions = sessions.filter(team__team_type=team_type)
        if team:
            sessions = sessions.filter(team=team)
        if start:
            sessions = sessions.filter(start_at__date__gte=start)
        if end:
            sessions = sessions.filter(start_at__date__lte=end)
    team_qs = Team.objects.filter(is_active=True).select_related("sport").order_by("sport__name", "name")
    if not is_admin_user(request.user):
        team_qs = team_qs.filter(coordinator=request.user)
    can_schedule_any = is_admin_user(request.user) or (role_for(request.user) in TRAINER_ROLES and team_qs.exists())
    session_list = add_session_permissions(request.user, sessions)
    venues = Venue.objects.filter(is_active=True).order_by("name")
    return render(request, "core/sessions.html", {
        "sessions": session_list,
        "form": form,
        "teams": team_qs,
        "venues": venues,
        "can_schedule_any": can_schedule_any,
        "session_stats": session_stats,
    })


@login_required
def meetings_list(request):
    preview = request.session.get("meeting_preview")
    can_schedule_meeting = is_admin_user(request.user)
    if request.method == "POST":
        if not can_schedule_meeting:
            messages.error(request, "You cannot schedule meetings.")
            return redirect("meetings")
        action = request.POST.get("action")
        meeting_id = request.POST.get("meeting_id")
        meeting = get_object_or_404(Meeting, pk=meeting_id) if meeting_id else None
        if action == "preview":
            built = build_meeting_preview(request.POST, organizer=request.user)
            if built["errors"]:
                for error in built["errors"]:
                    messages.error(request, error)
                request.session["meeting_draft"] = built["data"]
                request.session.modified = True
                return redirect("meetings")
            else:
                request.session["meeting_preview"] = built
                request.session.modified = True
                return redirect("meetings")
        elif action == "cancel_preview":
            built = request.session.get("meeting_preview")
            if built:
                request.session["meeting_draft"] = built["data"]
            request.session.pop("meeting_preview", None)
            return redirect("meetings")
        elif action == "confirm_create":
            built = request.session.get("meeting_preview")
            if not built:
                messages.error(request, "Meeting preview expired. Please schedule again.")
                return redirect("meetings")
            data = built["data"]
            preview_data = built.get("preview", {})
            sports = Sport.objects.filter(pk__in=data["sport_ids"], is_active=True)
            teams = Team.objects.filter(pk__in=data["team_ids"], is_active=True, sport__in=sports)
            trainers = User.objects.filter(pk__in=data["trainer_ids"], is_active=True)
            participants = User.objects.filter(pk__in=data["participant_ids"], is_active=True)
            if not participants.exists():
                messages.error(request, "No valid participants found.")
                request.session.pop("meeting_preview", None)
                return redirect("meetings")
            with transaction.atomic():
                meeting = Meeting.objects.create(
                    title=data["title"],
                    meeting_date=data["meeting_date"],
                    start_time=data["start_time"],
                    end_time=data["end_time"],
                    venue=data["venue"],
                    agenda=data["agenda"],
                    scheduled_by=request.user,
                )
                meeting.sports.set(sports)
                meeting.teams.set(teams)
                meeting.trainers.set(trainers)
                meeting.participants.set(participants)
                create_notifications(
                    participants,
                    "New meeting scheduled",
                    f"{meeting.title} scheduled on {preview_data.get('display_date') or data['meeting_date']} from {preview_data.get('display_start_time') or data['start_time']} to {preview_data.get('display_end_time') or data['end_time']} at {meeting.venue}.",
                    actor=request.user,
                    target_url=reverse("meetings"),
                    email_type="meeting_scheduled",
                    email_context={"details": meeting_email_details(
                        meeting.title,
                        data["meeting_date"],
                        data["start_time"],
                        data["end_time"],
                        meeting.venue,
                        meeting.agenda,
                        request.user,
                        actor_label="Scheduled By",
                    )},
                )
            request.session.pop("meeting_preview", None)
            messages.success(request, "Meeting scheduled and participants notified.")
            return redirect("meetings")
        elif action == "update" and meeting:
            previous_schedule = saved_meeting_schedule_tuple(meeting)
            previous_participant_ids = set(meeting.participants.values_list("pk", flat=True))
            built = build_meeting_preview(request.POST, organizer=request.user)
            if built["errors"]:
                for error in built["errors"]:
                    messages.error(request, error)
            else:
                data = built["data"]
                sports = Sport.objects.filter(pk__in=data["sport_ids"], is_active=True)
                teams = Team.objects.filter(pk__in=data["team_ids"], is_active=True, sport__in=sports)
                trainers = User.objects.filter(pk__in=data["trainer_ids"], is_active=True)
                participants = User.objects.filter(pk__in=data["participant_ids"], is_active=True)
                if not participants.exists():
                    messages.error(request, "No valid participants found.")
                else:
                    with transaction.atomic():
                        meeting.title = data["title"]
                        meeting.meeting_date = data["meeting_date"]
                        meeting.start_time = data["start_time"]
                        meeting.end_time = data["end_time"]
                        meeting.venue = data["venue"]
                        meeting.agenda = data["agenda"]
                        meeting.save(update_fields=["title", "meeting_date", "start_time", "end_time", "venue", "agenda", "updated_at"])
                        meeting.sports.set(sports)
                        meeting.teams.set(teams)
                        meeting.trainers.set(trainers)
                        meeting.participants.set(participants)
                    updated_schedule = (data["meeting_date"], data["start_time"], data["end_time"], data["venue"])
                    is_rescheduled = previous_schedule != updated_schedule
                    notify_ids = previous_participant_ids | set(participants.values_list("pk", flat=True))
                    create_notifications(
                        User.objects.filter(pk__in=notify_ids, is_active=True),
                        "Meeting rescheduled" if is_rescheduled else "Meeting updated",
                        f"{meeting.title} {'rescheduled' if is_rescheduled else 'updated'} for {meeting_display_date(meeting.meeting_date)} from {meeting_display_time(meeting.start_time)} to {meeting_display_time(meeting.end_time)} at {meeting.venue}.",
                        actor=request.user,
                        target_url=reverse("meetings"),
                        email_type="meeting_rescheduled" if is_rescheduled else "meeting_updated",
                        email_context={"details": meeting_email_details(
                            meeting.title,
                            meeting.meeting_date,
                            meeting.start_time,
                            meeting.end_time,
                            meeting.venue,
                            meeting.agenda,
                            request.user,
                        )},
                    )
                    messages.success(request, "Meeting updated and participants notified.")
                    return redirect("meetings")
        elif action == "mark_completed" and meeting:
            ending_note = request.POST.get("ending_note", "").strip()
            meeting.status = Meeting.Status.COMPLETED
            meeting.ending_note = ending_note
            meeting.save(update_fields=["status", "ending_note", "updated_at"])
            participants = list(meeting.participants.filter(is_active=True))
            create_notifications(
                participants,
                "Meeting completed",
                f"{meeting.title} marked as completed for {meeting_display_date(meeting.meeting_date)}.",
                actor=request.user,
                target_url=reverse("meetings"),
                email_type="meeting_completed",
                email_context={"details": meeting_email_details(
                    meeting.title,
                    meeting.meeting_date,
                    meeting.start_time,
                    meeting.end_time,
                    meeting.venue,
                    ending_note or meeting.agenda,
                    request.user,
                    actor_label="Completed By",
                )},
            )
            messages.success(request, "Meeting marked as completed and participants notified.")
            return redirect("meetings")
        elif action in {"cancel", "delete"} and meeting:
            cancelled_title = meeting.title
            meeting_date = meeting.meeting_date
            start_time = meeting.start_time
            end_time = meeting.end_time
            venue = meeting.venue
            agenda = meeting.agenda
            participants = list(meeting.participants.filter(is_active=True))
            ending_note = request.POST.get("ending_note", "").strip()
            meeting.status = Meeting.Status.CANCELLED
            meeting.ending_note = ending_note
            meeting.save(update_fields=["status", "ending_note", "updated_at"])
            create_notifications(
                participants,
                "Meeting cancelled",
                f"{cancelled_title} cancelled for {meeting_display_date(meeting_date)} at {venue}.",
                actor=request.user,
                target_url=reverse("meetings"),
                email_type="meeting_cancelled",
                email_context={"details": meeting_email_details(
                    cancelled_title,
                    meeting_date,
                    start_time,
                    end_time,
                    venue,
                    ending_note or agenda,
                    request.user,
                    actor_label="Cancelled By",
                )},
            )
            messages.success(request, "Meeting cancelled and participants notified.")
            return redirect("meetings")
    meeting_qs = visible_meetings(request.user).prefetch_related("sports", "teams", "trainers", "participants").select_related("scheduled_by")
    meeting_stats = {
        "total": meeting_qs.count(),
        "upcoming": meeting_qs.filter(meeting_date__gte=timezone.localdate(), status=Meeting.Status.SCHEDULED).count(),
        "completed": meeting_qs.filter(status=Meeting.Status.COMPLETED).count(),
        "cancelled": meeting_qs.filter(status=Meeting.Status.CANCELLED).count(),
    }
    meetings = list(meeting_qs)
    for meeting in meetings:
        meeting.sport_ids_csv = ",".join(str(sport.pk) for sport in meeting.sports.all())
        meeting.team_ids_csv = ",".join(str(team.pk) for team in meeting.teams.all())
        meeting.trainer_ids_csv = ",".join(str(trainer.pk) for trainer in meeting.trainers.all())
        participant_ids = {participant.pk for participant in meeting.participants.all()}
        team_modes = []
        for team in meeting.teams.all():
            active_member_ids = set(Membership.objects.filter(team=team, is_active=True).values_list("user_id", flat=True))
            mode = "ALL" if active_member_ids and active_member_ids.issubset(participant_ids) else "LEADS"
            team_modes.append(f"{team.pk}:{mode}")
        meeting.team_modes_csv = ",".join(team_modes)
    sports = Sport.objects.filter(is_active=True).order_by("name")
    teams = Team.objects.filter(is_active=True).select_related("sport").order_by("sport__name", "gender", "name")
    venues = Venue.objects.filter(is_active=True).order_by("name")
    trainers = User.objects.filter(
        is_active=True,
        profile__role__in=[UserProfile.Role.TRAINER, UserProfile.Role.COORDINATOR],
    ).select_related("profile").order_by("first_name", "last_name", "email")
    meeting_draft = request.session.pop("meeting_draft", None)
    return render(request, "core/meetings.html", {
        "meetings": meetings,
        "sports": sports,
        "teams": teams,
        "venues": venues,
        "trainers": trainers,
        "meeting_preview": preview,
        "meeting_draft": meeting_draft,
        "can_schedule_meeting": can_schedule_meeting,
        "meeting_stats": meeting_stats,
    })


@login_required
def session_form(request, pk=None):
    session = get_object_or_404(Session, pk=pk) if pk else None
    if session and not can_manage_team(request.user, session.team):
        messages.error(request, "You cannot edit this session.")
        return redirect("sessions")
    original_team = session.team if session else None
    original_schedule = (session.start_at, session.end_at, session.schedule_slot, session.full_day, session.team_id) if session else None
    form = SessionForm(request.POST or None, instance=session)
    if not is_admin_user(request.user):
        form.fields["team"].queryset = Team.objects.filter(coordinator=request.user)
    if request.method == "POST" and form.is_valid():
        obj = form.save(commit=False)
        if not can_schedule_session(request.user, obj.team):
            messages.error(request, "You cannot schedule for this team.")
            return redirect("sessions")
        is_new_session = obj.pk is None
        obj.scheduled_by = obj.scheduled_by or request.user
        obj.save()
        if is_new_session:
            notify_practice_session(
                obj,
                "New practice session",
                f"{obj.title or 'Practice session'} scheduled for {obj.team}.",
                actor=request.user,
                email_type="practice_scheduled",
            )
        else:
            updated_schedule = (obj.start_at, obj.end_at, obj.schedule_slot, obj.full_day, obj.team_id)
            event_title = "Practice session rescheduled" if original_schedule != updated_schedule else "Practice session updated"
            impacted_teams = [obj.team]
            if original_team and original_team.pk != obj.team_id:
                impacted_teams.append(original_team)
            notify_practice_session(
                obj,
                event_title,
                f"{obj.title or 'Practice session'} updated for {obj.team}.",
                actor=request.user,
                teams=impacted_teams,
                email_type="practice_rescheduled" if original_schedule != updated_schedule else "practice_updated",
            )
        messages.success(request, "Practice session saved.")
        return redirect("sessions")
    return render(request, "core/form.html", {"form": form, "title": "Practice Session", "back_url": reverse("sessions")})


@login_required
def delegate_attendance(request, pk):
    session = get_object_or_404(Session, pk=pk)
    if not can_manage_team(request.user, session.team):
        messages.error(request, "You cannot delegate attendance for this session.")
        return redirect("sessions")
    form = DelegateForm(request.POST or None, session=session)
    if request.method == "POST" and form.is_valid():
        current_delegate = AttendanceDelegate.objects.filter(session=session).select_related("assigned_to").first()
        previous_delegate = current_delegate.assigned_to if current_delegate else None
        reason = form.cleaned_data["reason"]
        if form.cleaned_data.get("remove_delegate"):
            if current_delegate:
                current_delegate.delete()
                AttendanceDelegateLog.objects.create(
                    session=session,
                    previous_delegate=previous_delegate,
                    new_delegate=None,
                    changed_by=request.user,
                    reason=reason,
                )
                messages.success(request, "Attendance delegate removed.")
            else:
                messages.info(request, "No Session Incharge is currently assigned.")
            return redirect("sessions")
        new_delegate = form.cleaned_data["assigned_to"]
        delegate, created = AttendanceDelegate.objects.update_or_create(
            session=session,
            defaults={"assigned_to": new_delegate, "assigned_by": request.user, "reason": reason},
        )
        if created or previous_delegate != new_delegate:
            AttendanceDelegateLog.objects.create(
                session=session,
                previous_delegate=previous_delegate,
                new_delegate=new_delegate,
                changed_by=request.user,
                reason=reason,
            )
            create_notifications(
                [delegate.assigned_to],
                "Session Incharge assigned",
                f"You have been assigned as Session Incharge for {session.title or 'Practice session'} - {session.team} on {timezone.localtime(session.start_at):%d %b %Y}.",
                actor=request.user,
                target_url=reverse("sessions"),
                sport=session.team.sport,
                team=session.team,
                session=session,
                include_actor=True,
                email_type="session_incharge",
                email_context={"details": [("Sport", session.team.sport.name), ("Team", f"{session.team.get_gender_display()} {session.team.name}"), ("Session", session.title), ("Date", timezone.localtime(session.start_at).strftime("%d %B %Y")), ("Assigned By", user_label(request.user))]},
            )
        messages.success(request, "Attendance delegate assigned.")
        return redirect("sessions")
    current_delegate = AttendanceDelegate.objects.filter(session=session).select_related("assigned_to").first()
    delegate_logs = AttendanceDelegateLog.objects.filter(session=session).select_related("previous_delegate", "new_delegate", "changed_by")[:10]
    return render(request, "core/form.html", {
        "form": form,
        "title": "Assign Session Incharge",
        "back_url": reverse("sessions"),
        "current_delegate": current_delegate,
        "delegate_logs": delegate_logs,
    })


@login_required
def cancel_attendance_lock(request, pk):
    session = get_object_or_404(Session, pk=pk)
    if request.method != "POST":
        return redirect("sessions")
    if not can_take_attendance(request.user, session):
        messages.error(request, "You cannot cancel the attendance lock for this session.")
        return redirect("sessions")
    if session.attendance_submitted:
        messages.warning(request, "Attendance has already been submitted for this session.")
        return redirect("attendance_detail", pk=session.pk)
    session.attendance_started_by = None
    session.attendance_started_by_role = ""
    session.attendance_started_at = None
    session.attendance_lock_expires_at = None
    session.save(update_fields=["attendance_started_by", "attendance_started_by_role", "attendance_started_at", "attendance_lock_expires_at", "updated_at"])
    messages.success(request, "Attendance lock cancelled. Another authorized user can take attendance now.")
    return redirect("sessions")


@login_required
def take_attendance(request, pk):
    session = get_object_or_404(Session.objects.select_related("team"), pk=pk)
    if not can_take_attendance(request.user, session):
        messages.error(request, "You cannot take attendance for this session.")
        return redirect("sessions")
    if session.attendance_submitted:
        messages.warning(request, "Attendance has already been submitted for this session.")
        return redirect("attendance_detail", pk=session.pk)
    now = timezone.now()
    with transaction.atomic():
        locked_session = Session.objects.select_for_update().get(pk=session.pk)
        lock_active = (
            locked_session.attendance_started_by
            and locked_session.attendance_started_by != request.user
            and locked_session.attendance_lock_expires_at
            and locked_session.attendance_lock_expires_at > now
        )
        if lock_active:
            messages.warning(
                request,
                f"Attendance is already being taken by {locked_session.attendance_started_by.get_full_name() or locked_session.attendance_started_by.username} ({locked_session.attendance_started_by_role}).",
            )
            return redirect("sessions")
        if not locked_session.attendance_started_by or locked_session.attendance_lock_expires_at is None or locked_session.attendance_lock_expires_at <= now:
            locked_session.attendance_started_by = request.user
            locked_session.attendance_started_by_role = role_label(request.user)
            locked_session.attendance_started_at = now
        locked_session.attendance_lock_expires_at = now + timedelta(minutes=30)
        locked_session.save(update_fields=["attendance_started_by", "attendance_started_by_role", "attendance_started_at", "attendance_lock_expires_at", "updated_at"])
        session = locked_session
    members = User.objects.filter(memberships__team=session.team, memberships__is_active=True).distinct().order_by("first_name", "last_name", "username")
    if request.method == "POST":
        valid_statuses = {value for value, _ in AttendanceRecord.Status.choices}
        with transaction.atomic():
            locked_session = Session.objects.select_for_update().get(pk=session.pk)
            if locked_session.attendance_started_by and locked_session.attendance_started_by != request.user:
                messages.warning(request, "Another authorized user is currently taking this attendance.")
                return redirect("sessions")
            for member in members:
                status = request.POST.get(f"status_{member.pk}", AttendanceRecord.Status.ABSENT)
                if status not in valid_statuses:
                    status = AttendanceRecord.Status.ABSENT
                remarks = request.POST.get(f"remarks_{member.pk}", "")
                AttendanceRecord.objects.update_or_create(
                    session=locked_session,
                    member=member,
                    defaults={"status": status, "remarks": remarks, "marked_by": request.user},
                )
            locked_session.attendance_submitted = True
            locked_session.submitted_by = request.user
            locked_session.submitted_by_role = role_label(request.user)
            locked_session.submitted_at = timezone.now()
            locked_session.attendance_started_by = None
            locked_session.attendance_started_by_role = ""
            locked_session.attendance_started_at = None
            locked_session.attendance_lock_expires_at = None
            locked_session.save(update_fields=["attendance_submitted", "submitted_by", "submitted_by_role", "submitted_at", "attendance_started_by", "attendance_started_by_role", "attendance_started_at", "attendance_lock_expires_at", "updated_at"])
        messages.success(request, "Attendance submitted.")
        return redirect("attendance_detail", pk=session.pk)
    return render(request, "core/take_attendance.html", {"session": session, "members": members, "statuses": AttendanceRecord.Status.choices, "can_cancel_lock": can_take_attendance(request.user, session)})


@login_required
def attendance_detail(request, pk, export_type=None):
    session = get_object_or_404(visible_sessions(request.user), pk=pk)
    records = session.attendance_records.select_related("member", "marked_by")
    if export_type == "excel":
        return export_excel(records)
    if export_type == "pdf":
        return export_pdf(records)
    return render(request, "core/attendance_detail.html", {"session": session, "records": records})


@login_required
@role_required(*ADMIN_ROLES)
def edit_attendance(request, pk):
    record = get_object_or_404(AttendanceRecord.objects.select_related("member", "session"), pk=pk)
    old_status = record.status
    form = AttendanceEditForm(request.POST or None, instance=record)
    if request.method == "POST" and form.is_valid():
        updated = form.save()
        if old_status != updated.status:
            AttendanceEditLog.objects.create(
                attendance_record=updated,
                edited_by=request.user,
                old_status=old_status,
                new_status=updated.status,
                reason=form.cleaned_data.get("reason") or "No reason provided",
            )
        messages.success(request, "Attendance updated with audit history.")
        return redirect("attendance_detail", pk=record.session.pk)
    student_name = record.member.get_full_name() or record.member.email or record.member.username
    return render(request, "core/form.html", {"form": form, "title": f"Edit Attendance - {student_name}", "back_url": reverse("attendance_detail", args=[record.session.pk])})


@login_required
def my_attendance(request):
    records = AttendanceRecord.objects.filter(member=request.user).select_related("session", "session__team", "session__team__sport")
    return render(request, "core/my_attendance.html", {"records": records, "percentage": attendance_percentage(request.user)})


@login_required
def feedback_list(request):
    received = Feedback.objects.filter(receiver=request.user).select_related("sender", "session")
    sent = Feedback.objects.filter(sender=request.user).select_related("receiver", "session")
    received.update(is_read=True)
    sent_groups_by_key = {}
    for item in sent:
        sent_at_key = timezone.localtime(item.created_at).strftime("%Y%m%d%H%M")
        key = (item.session_id, item.message, sent_at_key)
        receiver_name = user_label(item.receiver)
        if key not in sent_groups_by_key:
            sent_groups_by_key[key] = {
                "receivers": [],
                "created_at": item.created_at,
                "message": item.message,
                "session": item.session,
            }
        sent_groups_by_key[key]["receivers"].append(receiver_name)
    sent_groups = sorted(sent_groups_by_key.values(), key=lambda item: item["created_at"], reverse=True)
    for item in sent_groups:
        item["receiver_names"] = ", ".join(item["receivers"])
    return render(request, "core/feedback.html", {"received": received, "sent_groups": sent_groups})


def feedback_type_for(sender, receiver):
    if is_admin_user(sender):
        return Feedback.FeedbackType.ADMIN_TO_STUDENT
    if role_for(sender) in TRAINER_ROLES:
        return Feedback.FeedbackType.COORDINATOR_TO_STUDENT
    return Feedback.FeedbackType.STUDENT_TO_ADMIN


def feedback_session_queryset(user):
    qs = Session.objects.select_related("team", "team__sport").order_by("-start_at")
    if is_admin_user(user):
        return qs
    if role_for(user) in TRAINER_ROLES:
        return qs.filter(team__coordinator=user)
    assigned_teams = Team.objects.filter(
        Q(memberships__user=user, memberships__is_active=True)
        | Q(captain=user)
        | Q(vice_captain=user)
    ).distinct()
    return qs.filter(team__in=assigned_teams)


def feedback_student_queryset(user):
    student_roles = [UserProfile.Role.MEMBER, UserProfile.Role.CAPTAIN, UserProfile.Role.VICE_CAPTAIN]
    qs = User.objects.filter(is_active=True, profile__role__in=student_roles).select_related("profile").order_by("first_name", "last_name", "email")
    if is_admin_user(user):
        return qs.distinct()
    if role_for(user) in TRAINER_ROLES:
        return qs.filter(memberships__team__coordinator=user, memberships__is_active=True).distinct()
    return User.objects.none()


def feedback_trainer_queryset():
    return User.objects.filter(
        is_active=True,
        profile__role__in=[UserProfile.Role.TRAINER, UserProfile.Role.COORDINATOR],
    ).select_related("profile").order_by("first_name", "last_name", "email").distinct()


@login_required
def send_feedback(request):
    sessions = feedback_session_queryset(request.user)
    students = feedback_student_queryset(request.user)
    trainers = feedback_trainer_queryset()
    admins = admin_users()
    user_role = role_for(request.user)
    if is_admin_user(request.user):
        recipient_options = [("student", "Student"), ("trainer", "Trainer"), ("both", "Both Student and Trainer")]
    elif user_role in TRAINER_ROLES:
        recipient_options = [("admin", "Admin"), ("student", "Student"), ("both", "Both Admin and Student")]
    else:
        recipient_options = [("admin", "Admin"), ("trainer", "Trainer"), ("both", "Both Admin and Trainer")]

    if request.method == "POST":
        recipient_type = request.POST.get("recipient_type")
        session_id = request.POST.get("session")
        student_ids = parse_int_ids(request.POST.getlist("students") or request.POST.getlist("student"))
        trainer_ids = parse_int_ids(request.POST.getlist("trainers") or request.POST.getlist("trainer"))
        message_text = (request.POST.get("message") or "").strip()
        session = sessions.filter(pk=session_id).first() if session_id else None
        receivers = []
        errors = []
        if recipient_type not in {item[0] for item in recipient_options}:
            errors.append("Select whom to send feedback to.")
        if not session:
            errors.append("Select Session.")
        if not message_text:
            errors.append("Feedback message is required.")

        if is_admin_user(request.user):
            if recipient_type in {"student", "both"}:
                selected_students = list(students.filter(pk__in=student_ids)) if student_ids else []
                if selected_students:
                    receivers.extend(selected_students)
                else:
                    errors.append("Select Student.")
            if recipient_type in {"trainer", "both"}:
                selected_trainers = list(trainers.filter(pk__in=trainer_ids)) if trainer_ids else []
                if selected_trainers:
                    receivers.extend(selected_trainers)
                else:
                    errors.append("Select Trainer.")
        elif user_role in TRAINER_ROLES:
            if recipient_type in {"admin", "both"}:
                receivers.extend(admins)
            if recipient_type in {"student", "both"}:
                selected_students = list(students.filter(pk__in=student_ids)) if student_ids else []
                if selected_students:
                    receivers.extend(selected_students)
                else:
                    errors.append("Select Student.")
        else:
            if recipient_type in {"admin", "both"}:
                receivers.extend(admins)
            if recipient_type in {"trainer", "both"}:
                selected_trainers = list(trainers.filter(pk__in=trainer_ids)) if trainer_ids else []
                if not selected_trainers and session and session.team.coordinator:
                    selected_trainers = [session.team.coordinator]
                if selected_trainers:
                    receivers.extend(selected_trainers)
                else:
                    errors.append("No trainer is assigned to the selected session.")

        unique_receivers = []
        seen = set()
        for receiver in receivers:
            if receiver and receiver.pk != request.user.pk and receiver.pk not in seen:
                seen.add(receiver.pk)
                unique_receivers.append(receiver)
        if not unique_receivers:
            errors.append("No receiver is available for this feedback.")
        if errors:
            for error in errors:
                messages.error(request, error)
        else:
            for receiver in unique_receivers:
                Feedback.objects.create(
                    sender=request.user,
                    receiver=receiver,
                    session=session,
                    feedback_type=feedback_type_for(request.user, receiver),
                    message=message_text,
                )
            create_notifications(
                unique_receivers,
                "New feedback received",
                f"{request.user.get_full_name() or request.user.email} sent you private feedback.",
                actor=request.user,
                target_url=reverse("feedback"),
                session=session,
                team=session.team if session else None,
                sport=session.team.sport if session else None,
                email_type="feedback_received",
                email_context={"details": [("From", user_label(request.user)), ("Session", session.title if session else "General feedback"), ("Received", timezone.localtime().strftime("%d %B %Y, %I:%M %p"))]},
            )
            messages.success(request, f"Feedback sent to {len(unique_receivers)} recipient(s).")
            return redirect("feedback")
    return render(request, "core/send_feedback.html", {
        "title": "Send Feedback",
        "recipient_options": recipient_options,
        "sessions": sessions,
        "students": students,
        "trainers": trainers,
        "is_feedback_admin": is_admin_user(request.user),
        "is_feedback_trainer": user_role in TRAINER_ROLES,
        "is_feedback_student": not is_admin_user(request.user) and user_role not in TRAINER_ROLES,
        "selected_student_ids": {str(item) for item in parse_int_ids(request.POST.getlist("students") or request.POST.getlist("student"))},
        "selected_trainer_ids": {str(item) for item in parse_int_ids(request.POST.getlist("trainers") or request.POST.getlist("trainer"))},
    })


@login_required
@role_required(*ADMIN_ROLES)
def reports(request, export_type=None):
    form = ReportFilterForm(request.GET or None)
    report = build_report(form.cleaned_data if form.is_valid() else {})
    if export_type == "excel":
        return export_report_excel(report)
    if export_type == "pdf":
        return export_report_pdf(report)
    return render(request, "core/reports.html", {
        "form": form,
        "report": report,
        "report_filter_data": report_filter_dependency_data(),
    })


def report_filter_dependency_data():
    teams = Team.objects.select_related("sport").order_by("sport__name", "gender", "team_type", "name")
    sessions = Session.objects.select_related("team", "team__sport").order_by("-start_at")
    students = User.objects.filter(
        Q(memberships__is_active=True)
        | Q(captained_teams__isnull=False)
        | Q(vice_captained_teams__isnull=False),
        is_active=True,
    ).distinct().prefetch_related("memberships__team__sport").order_by("first_name", "last_name", "email")
    trainers = User.objects.filter(
        profile__role__in=[UserProfile.Role.TRAINER, UserProfile.Role.COORDINATOR],
        is_active=True,
    ).distinct().order_by("first_name", "last_name", "email")
    venues = Session.objects.exclude(venue="").order_by("venue").values_list("venue", flat=True).distinct()
    return {
        "teams": [
            {
                "id": team.pk,
                "label": str(team),
                "sport": team.sport_id,
                "gender": team.gender,
                "team_type": team.team_type,
            }
            for team in teams
        ],
        "sessions": [
            {
                "id": session.pk,
                "label": f"{session.title or 'Practice session'} - {session.team} ({session.start_at:%d %b %Y})",
                "team": session.team_id,
                "sport": session.team.sport_id,
                "gender": session.team.gender,
                "team_type": session.team.team_type,
            }
            for session in sessions
        ],
        "students": [
            {
                "id": student.pk,
                "label": user_label(student),
                "teams": list(student.memberships.filter(is_active=True).values_list("team_id", flat=True)),
                "sports": list(student.memberships.filter(is_active=True).values_list("team__sport_id", flat=True).distinct()),
                "genders": list(student.memberships.filter(is_active=True).values_list("team__gender", flat=True).distinct()),
                "team_types": list(student.memberships.filter(is_active=True).values_list("team__team_type", flat=True).distinct()),
            }
            for student in students
        ],
        "trainers": [
            {
                "id": trainer.pk,
                "label": user_label(trainer),
                "teams": list(trainer.coordinated_teams.values_list("pk", flat=True)),
                "sports": list(trainer.coordinated_teams.values_list("sport_id", flat=True).distinct()),
                "genders": list(trainer.coordinated_teams.values_list("gender", flat=True).distinct()),
                "team_types": list(trainer.coordinated_teams.values_list("team_type", flat=True).distinct()),
            }
            for trainer in trainers
        ],
        "venues": [{"value": venue, "label": venue} for venue in venues],
    }


def percent_value(part, total):
    return round((part / total) * 100, 1) if total else 0


def report_cell(value, badge=""):
    return {"value": value if value not in {None, ""} else "-", "badge": badge}


def status_badge(status):
    return {
        AttendanceRecord.Status.PRESENT: "success",
        AttendanceRecord.Status.ABSENT: "danger",
        AttendanceRecord.Status.LATE: "warning",
        AttendanceRecord.Status.EARLY_EXIT: "info",
        "COMPLETED": "success",
        "UPCOMING": "primary",
        "SCHEDULED": "secondary",
    }.get(status, "")


def report_filter_labels(cleaned, report_type):
    relevant = {
        "attendance": ["sport", "gender", "team_type", "team", "session", "trainer", "student", "attendance_status", "start_date", "end_date"],
        "sessions": ["sport", "gender", "team_type", "team", "trainer", "venue", "session_status", "start_date", "end_date"],
        "gender": ["sport", "gender", "team_type", "trainer", "start_date", "end_date"],
        "sports_teams": ["sport", "gender", "team_type", "team", "trainer", "start_date", "end_date"],
        "trainer": ["trainer", "sport", "gender", "team_type", "team", "start_date", "end_date"],
        "feedback": ["sport", "gender", "team_type", "team", "session", "trainer", "student", "start_date", "end_date"],
    }.get(report_type, [])
    labels = {
        "sport": "Sport",
        "gender": "Gender",
        "team_type": "Team Category",
        "team": "Team",
        "session": "Practice Session",
        "trainer": "Trainer",
        "student": "Student",
        "attendance_status": "Attendance Status",
        "venue": "Venue",
        "session_status": "Session Status",
        "start_date": "Start Date",
        "end_date": "End Date",
    }
    values = []
    for key in relevant:
        value = cleaned.get(key)
        if not value:
            continue
        display = value
        if key == "gender":
            display = dict(Team.TeamGender.choices).get(value, value)
        elif key == "team_type":
            display = dict(Team.TeamType.choices).get(value, value)
        elif key == "attendance_status":
            display = dict(AttendanceRecord.Status.choices).get(value, value)
        elif key == "session_status":
            display = dict(ReportFilterForm.SESSION_STATUS_CHOICES).get(value, value)
        elif key in {"start_date", "end_date"}:
            display = value.strftime("%d %b %Y")
        values.append({"label": labels[key], "value": display})
    return values


def apply_session_filters(sessions, cleaned):
    sport = cleaned.get("sport")
    gender = cleaned.get("gender")
    team_type = cleaned.get("team_type")
    team = cleaned.get("team")
    trainer = cleaned.get("trainer")
    venue = cleaned.get("venue")
    session_status = cleaned.get("session_status")
    start = cleaned.get("start_date")
    end = cleaned.get("end_date")
    today = timezone.localdate()
    if sport:
        sessions = sessions.filter(team__sport=sport)
    if gender:
        sessions = sessions.filter(team__gender=gender)
    if team_type:
        sessions = sessions.filter(team__team_type=team_type)
    if team:
        sessions = sessions.filter(team=team)
    if trainer:
        sessions = sessions.filter(team__coordinator=trainer)
    if venue:
        sessions = sessions.filter(venue=venue)
    if session_status == "COMPLETED":
        sessions = sessions.filter(attendance_submitted=True)
    elif session_status == "UPCOMING":
        sessions = sessions.filter(start_at__date__gte=today, attendance_submitted=False)
    elif session_status == "SCHEDULED":
        sessions = sessions.filter(attendance_submitted=False)
    if start:
        sessions = sessions.filter(start_at__date__gte=start)
    if end:
        sessions = sessions.filter(start_at__date__lte=end)
    return sessions


def apply_attendance_filters(records, cleaned):
    records = records.filter(session__in=apply_session_filters(Session.objects.all(), cleaned))
    if cleaned.get("session"):
        records = records.filter(session=cleaned["session"])
    if cleaned.get("student"):
        records = records.filter(member=cleaned["student"])
    if cleaned.get("attendance_status"):
        records = records.filter(status=cleaned["attendance_status"])
    return records


def report_base(title, report_type, cleaned):
    return {
        "type": report_type,
        "title": title,
        "university": "Christ (Deemed to be University), Pune Lavasa Campus",
        "system": "Sports Attendance System",
        "generated_at": timezone.localtime(),
        "filters": report_filter_labels(cleaned, report_type),
        "summary": [],
        "columns": [],
        "rows": [],
    }


def build_report(cleaned):
    report_type = cleaned.get("report_type") or "attendance"
    builders = {
        "attendance": build_attendance_report,
        "sessions": build_practice_session_report,
        "gender": build_gender_report,
        "sports_teams": build_sports_team_report,
        "trainer": build_trainer_report,
        "feedback": build_feedback_report,
    }
    return builders.get(report_type, build_attendance_report)(cleaned)


def build_attendance_report(cleaned):
    report = report_base("Attendance Report", "attendance", cleaned)
    records = apply_attendance_filters(
        AttendanceRecord.objects.select_related("member", "member__profile", "marked_by", "session", "session__team", "session__team__sport"),
        cleaned,
    ).order_by("-session__start_at", "member__first_name")
    total = records.count()
    present = records.filter(status=AttendanceRecord.Status.PRESENT).count()
    absent = records.filter(status=AttendanceRecord.Status.ABSENT).count()
    late = records.filter(status=AttendanceRecord.Status.LATE).count()
    early = records.filter(status=AttendanceRecord.Status.EARLY_EXIT).count()
    attended = present + late + early
    report["summary"] = [
        {"label": "Attendance %", "value": f"{percent_value(attended, total)}%"},
        {"label": "Records", "value": total},
        {"label": "Present", "value": present},
        {"label": "Absent", "value": absent},
        {"label": "Late", "value": late},
        {"label": "Early Exit", "value": early},
    ]
    report["columns"] = ["Date", "Session", "Sport", "Team", "Student", "Status", "Trainer", "Marked By", "Marked At"]
    for record in records[:500]:
        report["rows"].append([
            report_cell(record.session.start_at.strftime("%d %b %Y")),
            report_cell(record.session.title or "Practice session"),
            report_cell(record.session.team.sport.name),
            report_cell(f"{record.session.team.get_gender_display()} {record.session.team.name}"),
            report_cell(record.member.get_full_name() or record.member.username),
            report_cell(record.get_status_display(), status_badge(record.status)),
            report_cell(user_label(record.session.team.coordinator) if record.session.team.coordinator else "-"),
            report_cell(user_label(record.marked_by) if record.marked_by else "-"),
            report_cell(timezone.localtime(record.marked_at).strftime("%d %b %Y, %I:%M %p") if record.marked_at else "-"),
        ])
    return report


def build_practice_session_report(cleaned):
    report = report_base("Practice Session Report", "sessions", cleaned)
    sessions = apply_session_filters(
        Session.objects.select_related("team", "team__sport", "team__coordinator").prefetch_related("attendance_records"),
        cleaned,
    ).order_by("-start_at")
    total = sessions.count()
    completed = sessions.filter(attendance_submitted=True).count()
    upcoming = sessions.filter(start_at__date__gte=timezone.localdate(), attendance_submitted=False).count()
    submitted_records = AttendanceRecord.objects.filter(session__in=sessions)
    present_like = submitted_records.filter(status__in=[AttendanceRecord.Status.PRESENT, AttendanceRecord.Status.LATE, AttendanceRecord.Status.EARLY_EXIT]).count()
    report["summary"] = [
        {"label": "Sessions", "value": total},
        {"label": "Completed", "value": completed},
        {"label": "Upcoming", "value": upcoming},
        {"label": "Attendance %", "value": f"{percent_value(present_like, submitted_records.count())}%"},
    ]
    report["columns"] = ["Date", "Schedule", "Session", "Sport", "Team", "Trainer", "Venue", "Participants", "Present", "Absent", "Attendance %", "Status"]
    for session in sessions[:500]:
        records = list(session.attendance_records.all())
        present = len([record for record in records if record.status in [AttendanceRecord.Status.PRESENT, AttendanceRecord.Status.LATE, AttendanceRecord.Status.EARLY_EXIT]])
        absent = len([record for record in records if record.status == AttendanceRecord.Status.ABSENT])
        participant_count = Membership.objects.filter(team=session.team, is_active=True).count()
        status = "COMPLETED" if session.attendance_submitted else "UPCOMING" if session.start_at.date() >= timezone.localdate() else "SCHEDULED"
        report["rows"].append([
            report_cell(session.start_at.strftime("%d %b %Y")),
            report_cell(session.get_schedule_slot_display()),
            report_cell(session.title or "Practice session"),
            report_cell(session.team.sport.name),
            report_cell(f"{session.team.get_gender_display()} {session.team.name}"),
            report_cell(user_label(session.team.coordinator) if session.team.coordinator else "-"),
            report_cell(session.venue),
            report_cell(participant_count),
            report_cell(present),
            report_cell(absent),
            report_cell(f"{percent_value(present, len(records))}%"),
            report_cell(status.title(), status_badge(status)),
        ])
    return report


def build_gender_report(cleaned):
    report = report_base("Gender-wise Report", "gender", cleaned)
    teams = Team.objects.select_related("sport", "coordinator")
    sessions = apply_session_filters(Session.objects.select_related("team", "team__sport"), cleaned)
    if cleaned.get("sport"):
        teams = teams.filter(sport=cleaned["sport"])
    if cleaned.get("gender"):
        teams = teams.filter(gender=cleaned["gender"])
    if cleaned.get("team_type"):
        teams = teams.filter(team_type=cleaned["team_type"])
    if cleaned.get("trainer"):
        teams = teams.filter(coordinator=cleaned["trainer"])
    report["columns"] = ["Gender", "Sports", "Teams", "Members", "Attendance Records", "Present", "Absent", "Late", "Early Exit", "Attendance %"]
    total_members = 0
    total_records = 0
    total_attended = 0
    gender_choices = Team.TeamGender.choices
    if cleaned.get("gender"):
        gender_choices = [(cleaned["gender"], dict(Team.TeamGender.choices).get(cleaned["gender"], cleaned["gender"]))]
    for gender_value, gender_label in gender_choices:
        gender_teams = teams.filter(gender=gender_value)
        gender_sessions = sessions.filter(team__in=gender_teams)
        records = AttendanceRecord.objects.filter(session__in=gender_sessions)
        members = Membership.objects.filter(team__in=gender_teams, is_active=True).values("user").distinct().count()
        present = records.filter(status=AttendanceRecord.Status.PRESENT).count()
        absent = records.filter(status=AttendanceRecord.Status.ABSENT).count()
        late = records.filter(status=AttendanceRecord.Status.LATE).count()
        early = records.filter(status=AttendanceRecord.Status.EARLY_EXIT).count()
        attended = present + late + early
        total_members += members
        total_records += records.count()
        total_attended += attended
        report["rows"].append([
            report_cell(gender_label),
            report_cell(gender_teams.values("sport").distinct().count()),
            report_cell(gender_teams.count()),
            report_cell(members),
            report_cell(records.count()),
            report_cell(present),
            report_cell(absent),
            report_cell(late),
            report_cell(early),
            report_cell(f"{percent_value(attended, records.count())}%"),
        ])
    report["summary"] = [
        {"label": "Genders", "value": len(Team.TeamGender.choices)},
        {"label": "Members", "value": total_members},
        {"label": "Attendance Records", "value": total_records},
        {"label": "Attendance %", "value": f"{percent_value(total_attended, total_records)}%"},
    ]
    return report


def build_sports_team_report(cleaned):
    report = report_base("Sports & Team Report", "sports_teams", cleaned)
    teams = Team.objects.select_related("sport", "captain", "vice_captain", "coordinator")
    if cleaned.get("sport"):
        teams = teams.filter(sport=cleaned["sport"])
    if cleaned.get("gender"):
        teams = teams.filter(gender=cleaned["gender"])
    if cleaned.get("team_type"):
        teams = teams.filter(team_type=cleaned["team_type"])
    if cleaned.get("team"):
        teams = teams.filter(pk=cleaned["team"].pk)
    if cleaned.get("trainer"):
        teams = teams.filter(coordinator=cleaned["trainer"])
    teams = teams.order_by("sport__name", "gender", "team_type", "name")
    sessions = apply_session_filters(Session.objects.filter(team__in=teams), cleaned)
    records = AttendanceRecord.objects.filter(session__in=sessions)
    report["summary"] = [
        {"label": "Sports", "value": teams.values("sport").distinct().count()},
        {"label": "Teams", "value": teams.count()},
        {"label": "Members", "value": Membership.objects.filter(team__in=teams, is_active=True).values("user").distinct().count()},
        {"label": "Sessions", "value": sessions.count()},
        {"label": "Attendance %", "value": f"{percent_value(records.exclude(status=AttendanceRecord.Status.ABSENT).count(), records.count())}%"},
    ]
    report["columns"] = ["Sport", "Gender", "Team Category", "Team", "Members", "Trainer", "Captain", "Vice Captain", "Sessions", "Completed", "Attendance %", "Status"]
    for team in teams[:500]:
        team_sessions = sessions.filter(team=team)
        team_records = records.filter(session__team=team)
        report["rows"].append([
            report_cell(team.sport.name),
            report_cell(team.get_gender_display()),
            report_cell(team.get_team_type_display()),
            report_cell(team.name),
            report_cell(Membership.objects.filter(team=team, is_active=True).count()),
            report_cell(user_label(team.coordinator) if team.coordinator else "-"),
            report_cell(user_label(team.captain) if team.captain else "-"),
            report_cell(user_label(team.vice_captain) if team.vice_captain else "-"),
            report_cell(team_sessions.count()),
            report_cell(team_sessions.filter(attendance_submitted=True).count()),
            report_cell(f"{percent_value(team_records.exclude(status=AttendanceRecord.Status.ABSENT).count(), team_records.count())}%"),
            report_cell("Active" if team.is_active else "Inactive", "success" if team.is_active else "secondary"),
        ])
    return report


def build_trainer_report(cleaned):
    report = report_base("Trainer Activity Report", "trainer", cleaned)
    trainers = User.objects.filter(profile__role__in=[UserProfile.Role.TRAINER, UserProfile.Role.COORDINATOR]).distinct()
    if cleaned.get("trainer"):
        trainers = trainers.filter(pk=cleaned["trainer"].pk)
    report["columns"] = ["Trainer", "Assigned Sports", "Assigned Teams", "Sessions", "Completed", "Upcoming", "Students", "Attendance %"]
    totals = {"teams": 0, "sessions": 0, "completed": 0, "students": 0}
    for trainer in trainers.order_by("first_name", "last_name", "email"):
        teams = Team.objects.filter(coordinator=trainer)
        if cleaned.get("sport"):
            teams = teams.filter(sport=cleaned["sport"])
        if cleaned.get("gender"):
            teams = teams.filter(gender=cleaned["gender"])
        if cleaned.get("team_type"):
            teams = teams.filter(team_type=cleaned["team_type"])
        if cleaned.get("team"):
            teams = teams.filter(pk=cleaned["team"].pk)
        sessions = apply_session_filters(Session.objects.filter(team__in=teams), cleaned)
        records = AttendanceRecord.objects.filter(session__in=sessions)
        team_count = teams.count()
        session_count = sessions.count()
        completed = sessions.filter(attendance_submitted=True).count()
        students = Membership.objects.filter(team__in=teams, is_active=True).values("user").distinct().count()
        totals["teams"] += team_count
        totals["sessions"] += session_count
        totals["completed"] += completed
        totals["students"] += students
        if not team_count and (cleaned.get("sport") or cleaned.get("gender") or cleaned.get("team_type") or cleaned.get("team")):
            continue
        report["rows"].append([
            report_cell(user_label(trainer)),
            report_cell(", ".join(teams.values_list("sport__name", flat=True).distinct()) or "-"),
            report_cell(team_count),
            report_cell(session_count),
            report_cell(completed),
            report_cell(sessions.filter(start_at__date__gte=timezone.localdate(), attendance_submitted=False).count()),
            report_cell(students),
            report_cell(f"{percent_value(records.exclude(status=AttendanceRecord.Status.ABSENT).count(), records.count())}%"),
        ])
    report["summary"] = [
        {"label": "Trainers", "value": len(report["rows"])},
        {"label": "Assigned Teams", "value": totals["teams"]},
        {"label": "Sessions", "value": totals["sessions"]},
        {"label": "Completed Sessions", "value": totals["completed"]},
        {"label": "Students", "value": totals["students"]},
    ]
    return report


def build_feedback_report(cleaned):
    report = report_base("Feedback Report", "feedback", cleaned)
    feedback = Feedback.objects.select_related("sender", "receiver", "session", "session__team", "session__team__sport")
    session_qs = apply_session_filters(Session.objects.all(), cleaned)
    if cleaned.get("session"):
        feedback = feedback.filter(session=cleaned["session"])
    else:
        feedback = feedback.filter(Q(session__in=session_qs) | Q(session__isnull=True))
    if cleaned.get("student"):
        feedback = feedback.filter(Q(sender=cleaned["student"]) | Q(receiver=cleaned["student"]))
    if cleaned.get("trainer"):
        feedback = feedback.filter(Q(sender=cleaned["trainer"]) | Q(receiver=cleaned["trainer"]))
    if cleaned.get("start_date"):
        feedback = feedback.filter(created_at__date__gte=cleaned["start_date"])
    if cleaned.get("end_date"):
        feedback = feedback.filter(created_at__date__lte=cleaned["end_date"])
    total = feedback.count()
    report["summary"] = [
        {"label": "Feedback Items", "value": total},
        {"label": "Unread", "value": feedback.filter(is_read=False).count()},
        {"label": "Student to Admin", "value": feedback.filter(feedback_type=Feedback.FeedbackType.STUDENT_TO_ADMIN).count()},
        {"label": "Admin/Trainer to Student", "value": feedback.exclude(feedback_type=Feedback.FeedbackType.STUDENT_TO_ADMIN).count()},
    ]
    report["columns"] = ["Date", "Type", "Session", "Sport", "Team", "Sender", "Receiver", "Status"]
    for item in feedback.order_by("-created_at")[:500]:
        team = item.session.team if item.session else None
        report["rows"].append([
            report_cell(timezone.localtime(item.created_at).strftime("%d %b %Y, %I:%M %p")),
            report_cell(item.get_feedback_type_display()),
            report_cell(item.session.title if item.session else "-"),
            report_cell(team.sport.name if team else "-"),
            report_cell(f"{team.get_gender_display()} {team.name}" if team else "-"),
            report_cell(user_label(item.sender) if item.sender else "-"),
            report_cell(user_label(item.receiver) if item.receiver else "-"),
            report_cell("Read" if item.is_read else "Unread", "secondary" if item.is_read else "primary"),
        ])
    return report


def export_report_excel(report):
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    title_fill = PatternFill("solid", fgColor="003366")
    header_fill = PatternFill("solid", fgColor="0056B3")
    section_fill = PatternFill("solid", fgColor="E7EBF1")
    white_font = Font(color="FFFFFF", bold=True)
    bold_font = Font(bold=True)
    ws.append([report["university"]])
    ws.append([report["system"]])
    ws.append([report["title"]])
    ws.append(["Generated", timezone.localtime(report["generated_at"]).strftime("%d %b %Y, %I:%M %p")])
    ws.append([])
    ws.append(["Selected Filters"])
    if report["filters"]:
        for item in report["filters"]:
            ws.append([item["label"], str(item["value"])])
    else:
        ws.append(["All", "No filters selected"])
    ws.append([])
    ws.append(["Summary"])
    for item in report["summary"]:
        ws.append([item["label"], item["value"]])
    ws.append([])
    detail_header_row = ws.max_row + 1
    ws.append(report["columns"])
    for row in report["rows"]:
        ws.append([cell["value"] for cell in row])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(report["columns"]), 2))
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max(len(report["columns"]), 2))
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=max(len(report["columns"]), 2))
    for row_number in [1, 2, 3]:
        cell = ws.cell(row=row_number, column=1)
        cell.font = white_font if row_number == 1 else bold_font
        cell.fill = title_fill if row_number == 1 else section_fill
        cell.alignment = Alignment(horizontal="center")
    for row in ws.iter_rows(min_row=6, max_row=ws.max_row):
        if row[0].value in {"Selected Filters", "Summary"}:
            row[0].font = bold_font
            row[0].fill = section_fill
    for cell in ws[detail_header_row]:
        cell.fill = header_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center")
    for column_cells in ws.columns:
        width = min(max(len(str(cell.value or "")) for cell in column_cells) + 2, 34)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = width
    ws.freeze_panes = ws.cell(row=detail_header_row + 1, column=1)
    output = BytesIO()
    wb.save(output)
    response = HttpResponse(output.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="{report["type"]}-report.xlsx"'
    return response


def export_report_pdf(report):
    buffer = BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    left = 34
    right = width - 34
    page = 1

    def draw_footer():
        pdf.setFont("Helvetica", 7)
        pdf.drawString(left, 22, "Sports Attendance System")
        pdf.drawRightString(right, 22, f"Page {page}")

    def draw_header():
        y_pos = height - 36
        logo_path = settings.BASE_DIR / "static" / "img" / "christ-logo.jpeg"
        if logo_path.exists():
            pdf.drawImage(str(logo_path), left, y_pos - 58, width=102, height=50, preserveAspectRatio=True, mask="auto")
        else:
            pdf.setStrokeColorRGB(0, .2, .4)
            pdf.setFillColorRGB(.93, .96, 1)
            pdf.roundRect(left, y_pos - 58, 62, 54, 8, stroke=1, fill=1)
            pdf.setFillColorRGB(0, .2, .4)
            pdf.setFont("Helvetica-Bold", 14)
            pdf.drawCentredString(left + 31, y_pos - 36, "CU")
        pdf.setFont("Helvetica-Bold", 12)
        pdf.drawString(left + 116, y_pos - 14, report["university"])
        pdf.setFont("Helvetica-Bold", 10)
        pdf.drawString(left + 116, y_pos - 30, report["system"])
        pdf.setFont("Helvetica-Bold", 11)
        pdf.drawString(left + 116, y_pos - 48, report["title"])
        return y_pos - 78

    def new_page():
        nonlocal page
        draw_footer()
        pdf.showPage()
        page += 1
        return draw_header()

    y = draw_header()
    pdf.setFont("Helvetica", 8)
    pdf.drawString(left, y, f"Generated: {timezone.localtime(report['generated_at']):%d %b %Y, %I:%M %p}")
    y -= 14
    filter_text = ", ".join([f"{item['label']}: {item['value']}" for item in report["filters"]]) or "Filters: All"
    pdf.drawString(left, y, filter_text[:140])
    y -= 20
    card_width = (right - left - 18) / 4
    for index, item in enumerate(report["summary"][:8]):
        if index and index % 4 == 0:
            y -= 46
        x = left + (index % 4) * (card_width + 6)
        pdf.setFillColorRGB(.93, .96, 1)
        pdf.roundRect(x, y - 32, card_width, 28, 5, stroke=0, fill=1)
        pdf.setFillColorRGB(0, 0, 0)
        pdf.setFont("Helvetica", 6.5)
        pdf.drawString(x + 6, y - 13, str(item["label"])[:24])
        pdf.setFont("Helvetica-Bold", 9)
        pdf.drawString(x + 6, y - 26, str(item["value"])[:18])
    y -= 56 if len(report["summary"]) <= 4 else 102
    col_count = max(len(report["columns"]), 1)
    col_width = (right - left) / col_count
    pdf.setFillColorRGB(0, .2, .4)
    pdf.rect(left, y - 12, right - left, 14, stroke=0, fill=1)
    pdf.setFillColorRGB(1, 1, 1)
    pdf.setFont("Helvetica-Bold", 5.7 if col_count > 9 else 6.5)
    for index, column in enumerate(report["columns"]):
        pdf.drawString(left + (index * col_width) + 2, y - 8, str(column)[:18])
    y -= 18
    pdf.setFillColorRGB(0, 0, 0)
    pdf.setFont("Helvetica", 5.7 if col_count > 9 else 6.4)
    for row in report["rows"][:500]:
        if y < 42:
            y = new_page()
            pdf.setFillColorRGB(0, .2, .4)
            pdf.rect(left, y - 12, right - left, 14, stroke=0, fill=1)
            pdf.setFillColorRGB(1, 1, 1)
            pdf.setFont("Helvetica-Bold", 5.7 if col_count > 9 else 6.5)
            for index, column in enumerate(report["columns"]):
                pdf.drawString(left + (index * col_width) + 2, y - 8, str(column)[:18])
            y -= 18
            pdf.setFillColorRGB(0, 0, 0)
            pdf.setFont("Helvetica", 5.7 if col_count > 9 else 6.4)
        for index, cell in enumerate(row):
            pdf.drawString(left + (index * col_width) + 2, y, str(cell["value"])[:18])
        y -= 12
    draw_footer()
    pdf.save()
    response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{report["type"]}-report.pdf"'
    return response
